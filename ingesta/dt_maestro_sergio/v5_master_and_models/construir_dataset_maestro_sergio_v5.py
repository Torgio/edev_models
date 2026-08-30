"""
TFM Energia UCM — Estructura de tiempo oficial del equipo (espina horaria + dataset de modelado)

Este script construye las DOS estructuras de tiempo que todo el equipo debe compartir de cara a
la presentacion del miercoles, para que cada quien pueda hacer su propio EDA/modelado sobre
exactamente la misma base:

  A) ESPINA HORARIA (`construir_espina_horaria`) -- union por timestamp exacto de las tablas
     nucleo (entsoe_gen_data, entsoe_load_inter, esios_gen, load_inter, esios_forecast_da) +
     variables diarias (commodities, capacidad) difundidas sobre las 24h. Para EDA, correlaciones,
     graficos. NOTA (19-ago-2026): esios_load_inter se elimino, sustituida por `load_inter`
     (demanda + interconexiones consolidadas, ESIOS+ENTSO-E unificado).

  B) DATASET DIARIO (`construir_dataset_diario`) -- una fila por dia D, target = las 24 horas de
     precio de D+1 (nunca del propio D), solo features sin fuga de informacion, lags de datos
     reales en D-1/D-7, y el split cronologico train/validation/test ya fijado. Para entrenar y
     comparar modelos -- el mismo split para todos, si no, comparar resultados el miercoles no
     tiene sentido.

Que resuelve el dataset diario, en una frase por punto:
  1. Una fila por dia D, target = las 24 horas de precio de D+1 (nunca del propio D).
  2. Solo features sin fuga de informacion (catalogo ya verificado con el equipo).
  3. Lags de datos reales (demanda, eolica/bombeo, solar, precio) en D-1 y D-7.
  4. Filtra automaticamente los dias con target incompleto (cambio de hora, borde de la ventana).
  5. Split cronologico train/validation/test ya fijado -- mismo periodo de test para todos.

Uso:
    from construir_dataset_maestro import (
        construir_espina_horaria, construir_dataset_diario, dividir_train_val_test,
    )

    # FORMATO DE TRABAJO OFICIAL: el horario. El diario queda como baseline tabular opcional.
    horario = construir_dataset_horario()
    trh, vah, teh = dividir_train_val_test_horario(horario)   # corta por dia, nunca por fila

    espina = construir_espina_horaria()          # para EDA/correlaciones
    dataset = construir_dataset_diario()          # baseline tabular, 1 fila/dia, 24 targets
    train, val, test = dividir_train_val_test(dataset)

    # Tensores encoder/decoder para el Seq2Seq (past covariates / future covariates)
    X_hist, X_fut, X_est, y, fechas, meta = construir_tensores(ventana_dias=7)
    tr, va, te = dividir_tensores(fechas)

O directamente desde la terminal, para generar los CSV compartidos:
    python construir_dataset_maestro.py                    # HORARIO, las 2 variantes de NTC
        -> dataset_horario_sin_ntc_prev_v01.csv / .meta.json / .xlsx
        -> dataset_horario_sin_2020_v01.csv     / .meta.json / .xlsx
    python construir_dataset_maestro.py --variante sin_2020 # solo una variante
    python construir_dataset_maestro.py --tensores          # + tensores encoder/decoder (.npz)
    python construir_dataset_maestro.py --con-diario        # + el diario, como baseline tabular
    python construir_dataset_maestro.py --con-espina        # + espina horaria (~35 MB, lento)

Requisitos: pandas, psycopg2-binary, holidays (mismo credentials.json que el resto de ingesta/).


═══════════════════════════════════════════════════════════════════════════════════════════
 PENDIENTES  (23-ago-2026)
═══════════════════════════════════════════════════════════════════════════════════════════

 [1] CARGA HISTORICA DE PDBC  -- BLOQUEA la ventana de entrenamiento
     `esios_pdbc_gen` no cubre desde 2020-01-01, y como esta en EXIGIR_COBERTURA recorta el
     inicio de TODOS los datasets. El script imprime al arrancar desde que dia hay dato y
     cuantos se pierden.
     Accion: cargar el historico 2020 -> inicio actual (script equivalente a los de
     ingesta/historic_load/). Hasta entonces, dos salidas posibles segun lo que diga el aviso:
       - perdida pequeña  -> se deja como esta, el recorte no duele.
       - perdida > ~15%   -> EXIGIR_COBERTURA = [] y quitar PDBC de las piezas del dataset.
         Se recupera el historico completo a cambio de 102 columnas que solo entran laggeadas.
     Decidir con el numero delante, no antes.

 [2] AUTOCONSUMO_INI esta puesto a 2025-12-01 A OJO. Afinar con:
       SELECT date_trunc('day', datetime) AS dia, avg(ree_load - entsoe_load) AS brecha_mw
       FROM load_inter WHERE datetime >= '2025-11-15' AND datetime < '2026-01-15'
       GROUP BY 1 ORDER BY 1;
     El primer dia con brecha de centenares de MW (en vez de +-1) es el valor correcto.

 [3] TOPE_GAS_INI / TOPE_GAS_FIN (excepcion iberica) sin verificar contra BOE.

 [4] `dias_desde_cierre` sale CONSTANTE, o sea que `commodities` tiene dato todos los dias
     naturales, findes incluidos -- pero MIBGAS/TTF/EUA no cotizan sabados ni domingos. El
     relleno se hace en el pipeline de ingesta, antes de escribir en la tabla. Consecuencia:
     una fuente caida es indistinguible de una viva a nivel de BD, que es exactamente como se
     colo el hueco de 13 dias de gas_ttf. La deteccion tiene que vivir en el pipeline.

 [5] Verificar si ESIOS revisa historicamente TODAS las series de COLS_SEGURAS_FORECAST, no
     solo demanda_residual_prev_mw. Los `_history.py` cargaron valores YA REVISADOS, que no son
     los que existian al predecir. Pasar check_tables/verificar_revision_indicadores.py sobre
     las 13 columnas -- es la comprobacion de mas valor por tiempo invertido que queda.

 [6] Revisar `capinst_autoconsume_battery_mw` y `capdisp_fuel_mw`: salen constantes en 6 años y
     medio. En la bateria de autoconsumo eso apunta a columna vacia, no a serie plana.

 [7] Bloque ntc_*_prev_mw (indicadores 1844-1850) vacio hasta nov-2020 (~15% del train) y
     REDUNDANTE con ree_ntc_* de load_inter, que si esta completo. Candidato a eliminar.
"""

import sys
from pathlib import Path

import warnings

import numpy as np
import pandas as pd
import psycopg2

# pandas avisa en CADA read_sql de que psycopg2 no es un connectable de SQLAlchemy. Son ~20
# lineas por ejecucion que tapan los AVISOS que si importan. La combinacion funciona
# perfectamente (es la que usa toda la ingesta del proyecto), asi que se silencia ese warning
# concreto -- no un filtro global.
warnings.filterwarnings(
    "ignore",
    message="pandas only supports SQLAlchemy connectable",
    category=UserWarning,
)

# Localizacion de `ingesta/config.py` (23-ago-2026). Antes era un `parent.parent` fijo, que
# solo funcionaba si el script vivia exactamente en <repo>/modelos/. Al moverlo un nivel mas
# abajo (p.ej. modelos/dt_maestro_sergio/) fallaba con ModuleNotFoundError: No module named
# 'config'. Ahora se sube por el arbol hasta encontrar la carpeta ingesta, asi que da igual
# donde se coloque el fichero dentro del repo.
_RAIZ_INGESTA = None
for _dir in Path(__file__).resolve().parents:
    if (_dir / "ingesta" / "config.py").exists():
        _RAIZ_INGESTA = _dir / "ingesta"
        break
if _RAIZ_INGESTA is None:
    raise ImportError(
        "No encuentro ingesta/config.py subiendo desde " + str(Path(__file__).resolve())
        + ". Ejecuta el script desde dentro del repo tfm-energia."
    )
sys.path.append(str(_RAIZ_INGESTA))
from config import load_config

# ── Ventana de datos y fronteras del split (fijas para todo el equipo) ──────────────────────
DATASET_START = "2020-01-01"
DATASET_END = "2026-08-24"   # 23-ago-2026: la BD llega al 24-08, se dejaban 9 dias fuera del test

TRAIN_END = pd.Timestamp("2024-12-31").date()   # train: DATASET_START -> TRAIN_END
VAL_END = pd.Timestamp("2025-12-31").date()      # validation: TRAIN_END+1 -> VAL_END
# test: VAL_END+1 -> MODELO_END

# CORTE DE MODELADO -- decision del equipo 23-ago-2026: el modelo llega hasta final de julio.
# OJO a la distincion, no son lo mismo:
#   DATASET_END  = hasta donde se LEEN datos de la BD (24-ago).
#   MODELO_END   = ultimo dia PREDICHO, es decir el ultimo D+1 que aparece como target (31-jul).
# Se mantiene la ventana de lectura mas amplia a proposito: la fila D=30-jul necesita el precio
# del 31-jul como target, y los lags D-1/D-7 y el ffill de commodities necesitan margen por
# detras. Recortar DATASET_END en vez de filtrar las filas dejaria el ultimo dia sin target y
# degradaria los ultimos lags. Para mover el corte, tocar SOLO esta constante.
MODELO_END = pd.Timestamp("2026-07-31").date()   # ultimo dia predicho (D+1)

# Descartar las filas sin precio es lo correcto para ENTRENAR -- una fila sin target no
# ensena nada -- y lo contrario de lo que necesita PREDECIR: el precio de manana no existe
# todavia, es justo lo que se quiere predecir, asi que con el filtro puesto esa fila se cae
# siempre por mucho que se suba MODELO_END.
# Por defecto True, o sea el comportamiento de siempre. Solo lo baja a False
# `scripts/construir_matriz_produccion.py`, y solo mientras construye la matriz del dia.
EXIGIR_TARGET = True

# ── Inicio efectivo por cobertura de fuentes (23-ago-2026) ─────────────────────────────────
# El dataset no arranca en DATASET_START si alguna fuente EXIGIDA no llega hasta ahi. Se
# calcula el primer dia en que TODAS las tablas de esta lista tienen dato y se recorta ahi.
# Motivo: preferimos menos filas completas que muchas filas con un bloque entero en blanco.
#
# LEE EL AVISO QUE IMPRIME AL EJECUTAR antes de dar esto por bueno. Es una decision con precio:
# si `esios_pdbc_gen` empieza tarde, recortar TODO el dataset por ese bloque cuesta años de
# entrenamiento a cambio de 102 columnas que ademas solo entran laggeadas (el PDBC de D+1 sale
# de la misma casacion que el precio, asi que como feature contemporanea no vale). Si la
# perdida supera ~15% del train, casi seguro compensa mas la opcion contraria: dejar la lista
# vacia, mantener el historico completo y sacar PDBC del dataset hasta cargar su historia.
# Solo INFORMATIVO: imprime desde cuando tiene dato cada tabla, sin recortar nada. El recorte
# real se hace fila a fila con EXIGIR_BLOQUES (ver justo debajo).
EXIGIR_COBERTURA: list = []

# ── Filtro FILA A FILA por bloque incompleto (23-ago-2026) ─────────────────────────────────
# Prefijos de bloque que TIENEN que traer dato para que una fila entre en el dataset. Se
# descarta la fila cuyo bloque venga entero en blanco, no se mueve el inicio global.
#
# POR QUE FILA A FILA Y NO RECORTANDO EL INICIO. Son cosas distintas y la diferencia se nota:
#   - Recortar el inicio al primer dia con PDBC tira tambien los dias en que el resto de
#     fuentes SI tenian dato, y ademas no arregla nada si PDBC tiene huecos sueltos por el
#     medio (una carga historica fallida, un cron caido tres dias).
#   - Filtrar fila a fila quita exactamente los dias afectados, sean los primeros o del medio,
#     y conserva todo lo demas.
# El script imprime cuantas filas caen y en que fechas, que es lo que hace falta para saber si
# toca lanzar la carga historica de ese bloque (ver PENDIENTES [1]).
EXIGIR_BLOQUES = ["pdbc_"]

# ══ PODA DE COLUMNAS REDUNDANTES (23-ago-2026) ═══════════════════════════════════════════
# El dataset horario llego a 219 columnas para 1.778 dias de entrenamiento: una variable por
# cada diez observaciones. Eso es sobreajuste por construccion, y ademas buena parte era
# informacion DUPLICADA, no informacion nueva:
#
#   (a) `*_Dm6` es enteramente DERIVABLE de `*_Dm1`. Ambas son la misma serie horaria real
#       desplazada; teniendo la de D-1, la de D-6 sale reindexando por fecha. Guardar las dos
#       es escribir el mismo dato dos veces. 20 columnas.
#   (b) `es_esios_Dm1` y `es_esios_Dm6` son el propio `target_price` desplazado. 2 columnas.
#       Se conserva `es_esios_D` porque va al DECODER alineado por hora (al predecir la hora 14
#       de D+1, tener el precio de la hora 14 de D en esa misma posicion es un ancla directa
#       por paso de tiempo, y eso la secuencia no lo da igual de accesible).
#   (c) PDBC pasaba de 102 columnas (17 tecnologias x 3 estadisticos x 2 lags). El `min` y el
#       `max` de nuclear o biomasa son series planas, y el lag7d de un PROGRAMA repite lo que
#       ya dice el lag1d. Se queda la media de D-1 de las seis tecnologias que se mueven.
#   (d) `cap_baleares_prev_mw`: el target es el precio PENINSULAR.
#
# Lo importante: NO se pierde informacion. Las series reales horarias siguen enteras en las
# `*_Dm1`, y de ahi el notebook reconstruye la ventana continua de 168 h para el encoder. Lo
# que desaparece es la duplicacion.
# Registro de todo lo que se descarta, para la hoja EXCLUSIONES del Excel. Cada filtro anota
# aqui QUE se quita, POR QUE y con que dato de respaldo -- si no queda trazado, dentro de dos
# meses nadie recuerda por que falta una columna, y el tribunal lo pregunta.
REGISTRO_EXCLUSIONES: list = []


def _anotar(tipo: str, elemento: str, motivo: str, detalle: str = ""):
    REGISTRO_EXCLUSIONES.append({"tipo": tipo, "elemento": str(elemento),
                                 "motivo": motivo, "detalle": str(detalle)})


PODAR_REDUNDANTES = True

# Tecnologias de PDBC que conservan senal (el resto son planas o marginales en la peninsula)
PDBC_TECNOLOGIAS_UTILES = ["wind", "solar_pv", "hydro_ugh", "ccgt", "pumping_gen", "pumping_cons"]
PDBC_ESTADISTICOS = ["mean"]          # min/max aportan poco sobre un programa
PDBC_LAGS = ["lag1d"]                 # lag7d de un programa duplica al lag1d

# ══ VARIANTES POR EL BLOQUE ntc_*_prev_mw (23-ago-2026) ══════════════════════════════════
# Los indicadores ESIOS 1844-1850 (NTC prevista) no arrancan hasta noviembre de 2020, asi que
# dejan ~10 meses en blanco al principio de la serie -- en torno al 15% del train. Hay dos
# formas de resolverlo y son excluyentes, asi que se generan las DOS matrices y se comparan:
#
#   "sin_ntc_prev"  quita las 6 columnas y CONSERVA todos los dias.
#   "sin_2020"      conserva las columnas y DESCARTA los dias sin ellas.
#
# La comparacion es el resultado, no el tramite. Dato para interpretarla: `ree_ntc_impfr`,
# `ree_ntc_expfr`, etc. (de load_inter) llevan la MISMA capacidad de interconexion, estan en el
# decoder y tienen historico completo desde 2020. Es decir, "sin_2020" sacrifica ~330 dias de
# entrenamiento a cambio de informacion que el modelo ya recibe por otra via. Si aun asi gana,
# significa que las dos series NO son equivalentes y eso hay que investigarlo.
COLS_NTC_PREV = ["ree_ntc_impfr_prev", "ree_ntc_expfr_prev", "ree_ntc_imppt_prev",
                 "ree_ntc_exppt_prev", "ree_ntc_impma_prev", "ree_ntc_expma_prev"]
VARIANTES = ["sin_ntc_prev", "sin_2020"]

# ══ COLUMNAS EXCLUIDAS DEL DATASET (23-ago-2026) ═════════════════════════════════════════
# `dias_desde_cierre` se añadio para detectar fuentes de commodities caidas: cuenta los dias
# transcurridos desde el ultimo dato REAL, de forma que si empieza a subir (5, 8, 13...) es que
# la serie lleva tiempo muerta y el modelo esta comiendo un precio congelado.
#
# En la practica sale CONSTANTE a 0 en las tres ejecuciones: `commodities` tiene dato todos los
# dias naturales, findes incluidos, pese a que MIBGAS/TTF/EUA no cotizan sabado ni domingo. O
# sea que el relleno ya lo hace el pipeline de ingesta ANTES de escribir en la tabla, y cuando
# llega aqui la informacion de "esto es un dato repetido" ya se perdio.
#
# Consecuencia (anotada en PENDIENTES [4]): a nivel de BD, una fuente caida es indistinguible
# de una viva. Asi se colo el hueco de 13 dias de gas_ttf. La deteccion tiene que vivir en el
# pipeline, cruzando el ultimo dato real contra la hora de ejecucion del cron via pipeline_log.
#
# Aqui la columna solo mete una constante, asi que se excluye. El calculo se mantiene en
# `_features_dia_d` para que baste con vaciar esta lista si algun dia el pipeline deja de
# rellenar y la columna vuelve a tener sentido.
COLS_EXCLUIDAS = ["dias_desde_cierre"]

# `cap_baleares_prev_mw`: capacidad prevista del enlace con BALEARES. El target de este trabajo
# es el precio del sistema PENINSULAR, y el enlace balear no participa en su formacion -- es un
# sistema no peninsular con su propio despacho. Estaba en el script pero no en la matriz FINAL
# del equipo, asi que se retira (23-ago-2026). Poner a True para regenerar con ella y comparar.
# ══ PERIODOS EXCLUIDOS POR ANOMALIA (23-ago-2026) ════════════════════════════════════════
# Apagon iberico del 28-abr-2025 y reposicion posterior. Durante esos dias el precio no se
# formo por el mecanismo habitual: hubo cero de suministro, restricciones tecnicas masivas y un
# arranque escalonado del sistema. No es una observacion "extrema" que el modelo deba aprender,
# es una observacion de OTRO proceso -- mantenerla enseñaria al modelo relaciones que no
# volveran a darse, y ademas contaminaria las metricas de validacion.
#
# Caen dentro del periodo de VALIDACION (2025), asi que se pierden ~9 de 357 dias.
#
# CUIDADO AL USARLO EN EL NOTEBOOK: borrar dias del medio de la serie rompe la CONTIGUIDAD de
# la ventana del encoder. `PRECIO[t-V:t]` opera sobre los dias DISPONIBLES, asi que la fila del
# 7-may tomaria como "ultimos 7 dias" una mezcla que salta el hueco -- siete jornadas no
# consecutivas, sin que nada lo advierta. Hay que descartar tambien las muestras cuya ventana
# pise el hueco; el dataset marca esos dias en la columna `dia_excluido` para que el notebook
# pueda hacerlo (ver nota al final del constructor horario).
PERIODOS_EXCLUIDOS = [("2025-04-28", "2025-05-06")]   # apagon iberico + reposicion

# ══ CAPACIDAD COMO CUOTA, NO COMO NIVEL (24-ago-2026) ════════════════════════════════════
# Las columnas capinst_* son TENDENCIAS MONOTONAS: el parque solar y de baterias crece año a
# año y la cogeneracion se retira. Medido sobre el dataset generado, el desplazamiento entre
# train (<=2024) y test (2026) es brutal:
#
#     capinst_battery_hybrid_mw    +32,2 sigmas
#     capinst_cogeneration_mw      -11,4
#     capinst_hydro_mw             -10,1
#     capinst_other_renewable_mw   +10,0
#     ... 14 de 30 columnas por encima de 2 sigmas
#
# Al estandarizar con la media de train, esos valores caen SIEMPRE fuera del rango visto, y
# ningun modelo extrapola una tendencia: el arbol se queda en la ultima hoja aprendida y la red
# satura. Con el clip a +-10 sigmas la columna entra como una constante -- ruido, no señal.
# Se detecto porque al podar los estaticos de 183 a 46 estas columnas pasaron a dominar y el
# desplazamiento medio de X_est en test salto de +0,28 a +0,88 sigmas.
#
# SOLUCION: guardar la CUOTA en vez del nivel.
#   capinst_X / capinst_total_mw   -> que fraccion del parque es esa tecnologia. Estacionaria,
#       y ademas es lo que determina el orden de merito del despacho, que es el mecanismo por
#       el que la capacidad afecta al precio.
#   capdisp_X / capinst_X          -> que fraccion del parque de esa tecnologia esta
#       disponible. Es la magnitud que de verdad mueve el precio (indisponibilidades).
#
# Los MW absolutos siguen en la BD; aqui solo cambia lo que se escribe en el dataset.
CAPACIDAD_COMO_CUOTA = True

# ══ FILTRO DE TENDENCIAS MONOTONAS (24-ago-2026) ═════════════════════════════════════════
# La conversion a cuota NO basta. Medido: el desplazamiento maximo train->test solo baja de
# 32,2 a 27,8 sigmas, y siguen 14 columnas por encima de 3. El motivo es que el problema no es
# de ESCALA sino de EXISTENCIA: capinst_battery_hybrid_mw vale casi cero hasta 2024, asi que su
# desviacion tipica en train es minuscula y cualquier valor de 2026 queda a decenas de sigmas,
# en MW o en cuota da igual.
#
# El criterio correcto es detectar la tendencia SOBRE TRAIN, sin mirar test -- usar test para
# seleccionar variables seria contaminar la evaluacion. Una columna cuya correlacion de rango
# con el tiempo dentro de train supera el umbral es una tendencia monotona, y ningun modelo
# extrapola una tendencia: el arbol se queda en la ultima hoja aprendida y la red satura.
#
# LO QUE SE PIERDE, dicho sin adornos: capinst_solar_pv_mw es tendencia Y es informativa (mas
# solar instalada = precios de mediodia mas bajos). Al quitarla se pierde esa señal. Pero
# entraba al modelo como una constante clipada, asi que no la estaba aportando de todas formas.
# La via correcta para recuperarla seria la interaccion con la prevision solar horaria, que ya
# esta en el decoder -- queda anotado como mejora futura.
FILTRAR_TENDENCIAS = True
UMBRAL_TENDENCIA = 0.90        # |rho de Spearman con el tiempo| dentro de train
TRAIN_END_TENDENCIA = "2024-12-31"

# Segundo criterio, complementario. Las tecnologias HIBRIDAS y la bateria no tienen tendencia
# monotona -- valen cero durante casi todo train y luego arrancan -- asi que Spearman no las
# detecta, pero son las que peor se comportan: capinst_battery_hybrid_mw queda a +27,8 sigmas
# en test. El problema aqui no es la forma de la serie sino que la tecnologia NO EXISTIA
# mientras el modelo aprendia. Si una columna es nula o cero en mas de la mitad de train, el
# modelo no ha podido aprender nada sobre ella y en test solo aporta un valor fuera de rango.
COBERTURA_MINIMA_TRAIN = 0.50


def _filtrar_tendencias(ctx: pd.DataFrame) -> pd.DataFrame:
    """Descarta columnas con tendencia monotona fuerte, medida SOLO en train."""
    if not FILTRAR_TENDENCIAS or ctx.empty:
        return ctx
    idx = pd.to_datetime(pd.Series(ctx.index))
    tr = (idx <= TRAIN_END_TENDENCIA).to_numpy()
    if tr.sum() < 30:
        return ctx
    t = pd.Series(np.arange(tr.sum()))
    fuera = []
    for c in ctx.columns:
        v = ctx.loc[tr, c].reset_index(drop=True)
        if v.notna().sum() < 30 or v.nunique() <= 1:
            continue
        rho = v.corr(t, method="spearman")
        if pd.notna(rho) and abs(rho) > UMBRAL_TENDENCIA:
            fuera.append((c, round(float(rho), 3)))
    if fuera:
        print(f"[tendencias] {len(fuera)} columnas descartadas por tendencia monotona en train "
              f"(|rho|>{UMBRAL_TENDENCIA}):")
        for c, r in sorted(fuera, key=lambda x: -abs(x[1])):
            print(f"    {c:<34} rho {r:+.3f}")
            _anotar("columna", c, "Tendencia monotona en train",
                    f"rho Spearman con el tiempo = {r:+.3f} (umbral {UMBRAL_TENDENCIA}). "
                    f"Ningun modelo extrapola una tendencia: en test cae fuera del rango visto")
        ctx = ctx.drop(columns=[c for c, _ in fuera])

    # segundo criterio: tecnologia inexistente durante train (ver nota en COBERTURA_MINIMA_TRAIN)
    sin_datos = []
    for c in ctx.columns:
        v = ctx.loc[tr, c]
        cob = float(((v.notna()) & (v.abs() > 1e-9)).mean())
        if cob < COBERTURA_MINIMA_TRAIN:
            sin_datos.append((c, round(cob, 3)))
    if sin_datos:
        print(f"[cobertura] {len(sin_datos)} columnas descartadas: la tecnologia apenas existia "
              f"en train (<{COBERTURA_MINIMA_TRAIN:.0%} de filas con valor):")
        for c, cob in sorted(sin_datos, key=lambda x: x[1]):
            print(f"    {c:<34} cobertura {cob:.1%}")
            _anotar("columna", c, "Tecnologia inexistente en train",
                    f"solo {cob:.1%} de las filas de train tienen valor (umbral "
                    f"{COBERTURA_MINIMA_TRAIN:.0%}). El modelo no pudo aprender nada sobre ella")
        ctx = ctx.drop(columns=[c for c, _ in sin_datos])
    return ctx


def _normalizar_capacidad(ctx: pd.DataFrame) -> pd.DataFrame:
    """capinst_* -> cuota sobre el total; capdisp_* -> fraccion disponible del parque."""
    if not CAPACIDAD_COMO_CUOTA:
        return ctx
    # ORDEN IMPORTANTE: primero capdisp/capinst (ambas en MW), y solo despues convertir
    # capinst a cuota. Al reves, se dividiria un MW entre una fraccion.
    n_disp = 0
    for c in [c for c in ctx.columns if c.startswith("capdisp_")]:
        base = "capinst_" + c[len("capdisp_"):]
        # coal_antracita no tiene homologo exacto en instalada
        if base not in ctx.columns:
            base = {"capdisp_coal_antracita_mw": "capinst_coal_mw"}.get(c)
        if base and base in ctx.columns:
            ctx[c] = ctx[c] / ctx[base].replace(0, np.nan)
            n_disp += 1

    n_inst = 0
    if "capinst_total_mw" in ctx.columns:
        tot = ctx["capinst_total_mw"].replace(0, np.nan)
        for c in [c for c in ctx.columns
                  if c.startswith("capinst_") and c != "capinst_total_mw"]:
            ctx[c] = ctx[c] / tot
            n_inst += 1
        # el total en MW es la columna con mas riesgo de extrapolacion de todas: fuera
        ctx = ctx.drop(columns=["capinst_total_mw"])

    print(f"[capacidad] {n_inst} capinst_* -> cuota sobre el total | "
          f"{n_disp} capdisp_* -> fraccion disponible del parque")
    if n_inst:
        _anotar("transformacion", "capinst_*", "Convertida a cuota sobre capinst_total_mw",
                f"{n_inst} columnas. El nivel en MW es una tendencia monotona; la cuota es "
                f"estacionaria y es lo que fija el orden de merito del despacho")
        _anotar("columna", "capinst_total_mw", "Eliminada tras normalizar",
                "es la columna con mas riesgo de extrapolacion de todo el bloque")
    if n_disp:
        _anotar("transformacion", "capdisp_*", "Convertida a fraccion disponible del parque",
                f"{n_disp} columnas, divididas por su capinst_* homologa")
    return ctx

INCLUIR_CAP_BALEARES = False
if INCLUIR_CAP_BALEARES:
    COLS_SEGURAS_FORECAST_EXTRA = ["cap_baleares_prev_mw"]
else:
    COLS_SEGURAS_FORECAST_EXTRA = []


def _filtrar_filas_incompletas(df: pd.DataFrame, prefijos: list = None,
                               col_fecha: str = None) -> pd.DataFrame:
    """Descarta filas cuyo bloque exigido venga entero en blanco."""
    prefijos = EXIGIR_BLOQUES if prefijos is None else prefijos
    if not prefijos:
        return df
    for pref in prefijos:
        cols = [c for c in df.columns if c.startswith(pref)]
        if not cols:
            print(f"AVISO [filas]: no hay ninguna columna `{pref}*`; no se filtra por ese bloque")
            continue
        hay = df[cols].notna().any(axis=1)
        n = int((~hay).sum())
        if n:
            fechas = df.loc[~hay, col_fecha] if col_fecha else pd.Series(df.index[~hay])
            fechas = pd.to_datetime(fechas)
            print(f"Descartadas {n} filas sin datos de `{pref}*` "
                  f"({fechas.dt.date.nunique()} dias distintos, "
                  f"de {fechas.min().date()} a {fechas.max().date()})")
            _anotar("fechas", f"{fechas.min().date()} -> {fechas.max().date()}",
                    f"Sin datos del bloque `{pref}*`",
                    f"{n} filas / {fechas.dt.date.nunique()} dias. Se exige ese bloque "
                    f"completo (EXIGIR_BLOQUES)")
            # si son consecutivos desde el principio es un arranque tardio de la fuente;
            # si estan repartidos, son huecos de ingesta -- se distingue mirando el rango
            df = df[hay]
    return df

# ── Filtro de columnas por cobertura (23-ago-2026) ─────────────────────────────────────────
# Una columna que esta vacia la mitad de la ventana no aporta: mete NaN, obliga a imputar y el
# modelo aprende a leer la imputacion en vez del fenomeno. Se descartan las que no lleguen a
# COBERTURA_MINIMA... con UNA EXCEPCION IMPORTANTE.
#
# NO TODO NULO INICIAL SIGNIFICA "FALTA EL DATO". Hay dos casos distintos y confundirlos cuesta
# caro en este proyecto:
#
#   (a) NULO ESTRUCTURAL -> la tecnologia NO EXISTIA. ree_gbattery_mw esta vacia en 2020 porque
#       no habia baterias conectadas a la red. Ahi el valor fisicamente correcto es CERO, no
#       "desconocido". Descartar esas columnas seria tirar justo la tecnologia sobre la que va
#       el TFM. Se rellenan con 0 y se conservan.
#
#   (b) NULO DE FUENTE -> el dato existia pero no se publicaba o no se ingesto. Los
#       ntc_*_prev_mw (indicadores 1844-1850) no arrancan hasta nov-2020, y ademas son
#       REDUNDANTES con ree_ntc_* de load_inter, que si esta completo desde el principio. Esas
#       si se descartan.
#
# Es el mismo criterio COALESCE de siempre: decidir explicitamente si la ausencia propaga
# incertidumbre (NULL) o afirma que no hay flujo (0).
COBERTURA_MINIMA = 0.80          # fraccion de filas con dato exigida a cada columna

# Patrones de nombre cuyo nulo se interpreta como CERO ESTRUCTURAL (caso (a)).
# 24-ago-2026: se retiran "battery"/"bateria" de la lista. Los indicadores ESIOS 2166/2167 no
# publican hasta el 20-nov-2024 y train acaba el 31-dic-2024, asi que la cobertura en train es
# del 24,2%: rellenarlas a 0 hace que el modelo aprenda "la bateria siempre vale 0" y luego se
# encuentre cientos de MW en test. Es el mismo patron que eliminan los filtros de tendencia.
# Que caigan por cobertura, como las hibridas.
# LIMITACION PARA LA MEMORIA: con este split no se puede modelar el efecto de la bateria sobre
# el precio. Refuerza el planteamiento del capitulo de optimizacion, donde la bateria es
# variable de DECISION y no feature explicativa.
CEROS_ESTRUCTURALES = ["hybrid", "hibrid", "autoconsume", "autoconsumo"]


def _es_cero_estructural(col: str) -> bool:
    c = col.lower()
    return any(pat in c for pat in CEROS_ESTRUCTURALES)


def _filtrar_por_cobertura(df: pd.DataFrame, proteger: list = None,
                           minimo: float = None) -> pd.DataFrame:
    """Descarta columnas con cobertura insuficiente; rellena a 0 los ceros estructurales.

    `proteger` son columnas que nunca se tocan (claves, target, split). Imprime una tabla con
    nombre, cobertura y PRIMER DIA CON DATO de todo lo que descarta, que es la informacion que
    hace falta para decidir si toca cargar historico o eliminar la columna del catalogo.
    """
    minimo = COBERTURA_MINIMA if minimo is None else minimo
    proteger = set(proteger or [])
    cob = df.notna().mean()

    vacias, estructurales, pobres = [], [], []
    for c in df.columns:
        if c in proteger:
            continue
        if cob[c] == 0:
            vacias.append(c)
        elif cob[c] < minimo:
            (estructurales if _es_cero_estructural(c) else pobres).append(c)

    def _primer_dato(c):
        """Primera FECHA con dato. En el dataset horario el indice es un RangeIndex, asi que
        hay que usar la columna fecha_objetivo -- si no, se imprimia la posicion de la fila
        (p.ej. '42856') en lugar del dia, y la hoja EXCLUSIONES quedaba inservible."""
        m = df[c].notna()
        if not m.any():
            return "-"
        if isinstance(df.index, pd.RangeIndex) and "fecha_objetivo" in df.columns:
            return pd.to_datetime(df.loc[m, "fecha_objetivo"]).min().date()
        return pd.Series(df.index)[m.to_numpy()].min()

    if vacias:
        print(f"\nDescartadas {len(vacias)} columnas COMPLETAMENTE VACIAS: {vacias}")
        for c in vacias:
            _anotar("columna", c, "Completamente vacia", "0% de filas con valor")
    if estructurales:
        print(f"\nRellenadas a 0 ({len(estructurales)} columnas, nulo estructural = tecnologia "
              f"aun no conectada, no dato perdido):")
        for c in sorted(estructurales):
            print(f"  {c:<42} cobertura {cob[c]:5.1%}  primer dato {_primer_dato(c)}")
        df[estructurales] = df[estructurales].fillna(0)
    if pobres:
        print(f"\nDescartadas {len(pobres)} columnas por cobertura < {minimo:.0%} "
              f"(la fuente no publicaba; revisar si toca carga historica o quitar del catalogo):")
        for c in sorted(pobres):
            print(f"  {c:<42} cobertura {cob[c]:5.1%}  primer dato {_primer_dato(c)}")

    return df.drop(columns=vacias + pobres)


def _inicio_efectivo(conn, tablas: list = None) -> str:
    """Primer dia en que todas las tablas exigidas tienen dato. Devuelve fecha ISO."""
    tablas = EXIGIR_COBERTURA if tablas is None else tablas
    inicio = pd.Timestamp(DATASET_START)
    detalle = []
    for t in tablas:
        try:
            ts = _col_temporal(conn, t)
            m = pd.read_sql(f"SELECT min({ts}) AS m FROM {t}", conn)["m"].iloc[0]
        except Exception as e:
            print(f"AVISO [cobertura]: no puedo leer `{t}` ({e}); se ignora")
            continue
        if m is None:
            print(f"AVISO [cobertura]: `{t}` esta VACIA; se ignora")
            continue
        m = pd.Timestamp(pd.to_datetime(m).date())
        detalle.append((t, m.date()))
        inicio = max(inicio, m)

    if detalle:
        print("Cobertura de fuentes exigidas:")
        for t, d in detalle:
            print(f"  {t:<28} desde {d}")
    perdidos = (inicio - pd.Timestamp(DATASET_START)).days
    if perdidos > 0:
        total = (pd.Timestamp(MODELO_END) - pd.Timestamp(DATASET_START)).days
        pct = 100 * perdidos / total
        print(f"INICIO EFECTIVO: {inicio.date()} (en vez de {DATASET_START})")
        print(f"  -> se descartan {perdidos} dias, {pct:.1f}% de la ventana total")
        if pct > 15:
            print("  *** REVISAR: la perdida es grande. Valora dejar EXIGIR_COBERTURA = [] y")
            print("      sacar el bloque que llega tarde, en vez de recortar todo el historico. ***")
    return inicio.strftime("%Y-%m-%d")

# Features seguras de esios_forecast_da -- ampliado 20-ago-2026 segun la seleccion del equipo
# (matriz FORECAST). demanda_prev_mw (1775) SE MANTIENE junto a demanda_mercado_prev_mw (2563)
# a peticion del equipo, aunque el 1775 tiene un sesgo creciente documentado (hasta +2.386 MW en
# 2026, el doble de error todos los años frente al 2563) -- no es fuga, es redundancia con una
# version peor, se deja que el modelo decida. gen_renovables_prev_mw tambien se mantiene pese a
# ser colineal exacto con gen_wind_prev_mw + gen_solar_pv_prev_mw (verificado 18-ago-2026, 58.127
# filas identicas) -- mismo criterio, redundancia aceptada a proposito, no invalida el dataset.
COLS_SEGURAS_FORECAST = [
    "demanda_prev_mw", "demanda_mercado_prev_mw", "gen_wind_prev_mw", "gen_solar_pv_prev_mw",
    "gen_renovables_prev_mw", "gen_solartermica_prev_mw",
    "ntc_fr_imp_prev_mw", "ntc_fr_exp_prev_mw", "ntc_pt_imp_prev_mw", "ntc_pt_exp_prev_mw",
    "ntc_ma_imp_prev_mw", "ntc_ma_exp_prev_mw",
]
# AÑADIDAS 23-ago-2026 -- estaban en la matriz FINAL del equipo y faltaban en el script.
#   c_autoconsumo_prev   = 1775 - 2563. El modelo tiene las dos series y "podria" derivarla,
#     pero una red aprende muy mal una resta entre dos magnitudes grandes y casi iguales;
#     dandosela hecha el gradiente la ve directamente. r=0,957 con solar.
#   autoconsumo_estimado = flag booleano desde dic-2025. Es el marcador del cambio metodologico
#     de REE (seccion 3.2 del documento de metodologia), que cae dentro de validation/test pero
#     NO de train. Su media diaria funciona como dummy de regimen. Se castea a int en el SELECT
#     porque un boolean de postgres llega como bool de python y rompe agg(["mean"]).
# Se listan varios alias porque el Excel y la BD no coinciden en la convencion de nombres; el
# filtro por information_schema se queda con los que existan y avisa del resto.
# 23-ago-2026, comprobado contra la BD: NINGUNA de las dos existe en `esios_forecast_da`.
# La matriz las lista pero la tabla no las tiene. No hace falta tocar la BD: las dos se
# reconstruyen con lo que ya se lee.
#
#   c_autoconsumo_prev  = 1775 - 2563 = demanda_prev_mw - demanda_mercado_prev_mw.
#     Es la definicion literal de la columna del Excel, y las dos series ya estan en
#     COLS_SEGURAS_FORECAST. Se calcula en `_features_forecast`. Merece entrar explicita aunque
#     sea derivada: una red aprende muy mal una resta entre dos magnitudes grandes y casi
#     iguales, y esta diferencia correlaciona r=0,957 con la solar.
#
#   autoconsumo_estimado = flag de regimen que solo depende de la fecha (antes/despues de que
#     REE incorporase la estimacion de autoconsumo). No es un dato que haya que leer, es un
#     dummy de calendario -- ver AUTOCONSUMO_INI y `_calendario`.
COLS_FORECAST_AUTOCONSUMO: list = []
COL_FORECAST_FLAG = None
# EXCLUIDA -- OJO, NO es una omision, es una fuga real: demanda_residual_prev_mw se revisa
# 10-14 dias despues de publicarse (verificado con check_tables/verificar_revision_indicadores.py).
# El valor guardado hoy en la BD para una fecha historica es el YA REVISADO, no el que existia en
# el momento de predecir -- misma familia de bug que el del target D+1 que se corrigio el
# 17-ago-2026. Pendiente de confirmacion explicita antes de añadirla (ver conversacion 20-ago).
# potencia_indisp_pbf_mw ya NO EXISTE en esios_forecast_da (el equipo la elimino el 19-ago-2026 --
# era un duplicado exacto de esios_pbf_gen.unavailable_power_mw, y estar en la tabla de forecasts
# inducia a creerla leak-safe pre-cierre cuando en realidad es dato de PBF, post-cierre).

# Matriz de generacion/demanda del equipo, cerrada 19-ago-2026 (ver docs/notas_memoria_tfm.md y
# la vista materializada `generation` en la BD) -- reemplaza la seleccion anterior del 19-ago.
#
# Demanda real -> ree_load (load_inter), a peticion expresa del equipo (20-ago-2026), pese a que
# la version anterior de este script usaba entsoe_load por la contaminacion de autoconsumo
# documentada en D-03 (brecha +435 a +2.495 MW dic-2025/jul-2026). El equipo decidio usar ree_load
# de todas formas -- si se quiere volver a entsoe_load, es cambiar esta lista.
#
# entsoe_load añadida el 20-ago-2026 (adicion de ultimo momento del equipo, no un reemplazo de
# ree_load -- las dos entran juntas). OJO para quien la use: las dos miden lo mismo salvo por el
# autoconsumo (D-03), asi que en 2020-2024 son practicamente identicas y desde dic-2025 se separan
# cada vez mas -- entran ambas tal como se indico, sin resolver esa redundancia aqui.
#
# load_inter tambien aporta netflow/total_net_flow/gen_peninsular: son dato REAL ya ocurrido
# (no una capacidad publicada de antemano como el NTC), asi que van aqui con el resto de reales
# con lag D-1/D-7, no en COLS_NTC (que sigue el criterio "seguro, D+1" de _features_ntc).
TABLA_DEMANDA_REAL = "load_inter"
# ══ UNA SOLA SERIE DE DEMANDA (24-ago-2026) ══════════════════════════════════════════════
# `ree_load` (ESIOS 1293) sale del dataset. Motivo, tomado de la documentacion del equipo sobre
# load_inter, que ya lo tenia decidido -- "NO SE USA COMO FEATURE":
#
#   1. Hasta nov-2025 las dos series son la MISMA: medido, ree_load - entsoe_load = +-4 MW.
#      En practicamente todo el train son duplicados exactos.
#   2. Desde dic-2025 divergen hasta 2.495 MW porque ree_load incorpora la estimacion de
#      autoconsumo y entsoe_load no. Es decir: la diferencia entre ambas ES el autoconsumo que
#      acabamos de excluir -- mantener las dos lo reintroduciria por la puerta de atras.
#   3. El salto de ree_load en dic-2025 es de 1.582 MW, y el de la solar FV de 1.519 MW en el
#      mismo mes; mes a mes van de la mano (435/436 en dic, 686/688 en ene, 1.038/1.044 en feb).
#      El autoconsumo se genera y se consume en el mismo punto, asi que entra por los dos lados
#      del balance.
#
# Se conserva `entsoe_load` porque es METODOLOGICAMENTE ESTABLE en todo el historico, mientras
# que ree_load cambia de perimetro a mitad de la serie. Para un modelo entrenado hasta 2024, la
# serie que no cambia de definicion es la unica utilizable. Ademas es la validada con cinco
# pruebas independientes de balance del sistema (residuo < 2% sin estructura horaria).
#
# MATIZ para la memoria: entsoe_load incluye el consumo de bombeo. El art. 2.27 del Reglamento
# 543/2013 lo excluye del Actual Total Load, pero la implementacion española lo incluye --
# divergencia entre norma e implementacion, verificada empiricamente con el balance.
#
# CRITERIO GENERAL derivado: cualquier serie de ESIOS con autoconsumo dentro rompe en dic-2025.
# Aplica tambien a ree_gsolar_mw, al indicador 1775 y a los agregados 10351/10352.
DEMANDA_REAL = "entsoe_load"        # "entsoe_load" | "ree_load" | "ambas"

_DEM = {"entsoe_load": ["entsoe_load"], "ree_load": ["ree_load"],
        "ambas": ["ree_load", "entsoe_load"]}[DEMANDA_REAL]

# total_net_flow_mw y gen_peninsular_mw son GENERATED sobre columnas que ya estan aqui, asi que
# son redundantes en informacion. Se MANTIENEN por decision del equipo: gen_peninsular_mw tiene
# significado fisico propio (cuanta generacion peninsular hizo falta) y el modelo tendria que
# aprender a componerla. Ninguna lleva COALESCE, asi que propagan NULL en el apagon en vez de
# fabricar un cero falso.
COLS_DEMANDA_REAL = _DEM + ["ree_netflow_fr", "ree_netflow_pt", "ree_netflow_ma",
                            "total_net_flow_mw", "gen_peninsular_mw"]

# ENTSO-E gana eolica, hidraulica fluyente, consumo de bombeo, biomasa, residuos, otras
# renovables y fuel-oil. hydro_reservoir_mw y pumping_gen_mw YA NO van sueltas -- ver
# _features_lag_reales: se fusionan en hydro_dispatch_mw (equivalente a c_ghydrodispatch de la
# matriz) porque ENTSO-E no separaba la turbinacion de bombeo (B10) del embalse (B12) antes de
# dic-2022 -- pumping_gen_mw es NULL el 100% de 2020-2021 y 91% de 2022 (verificado 20-ago-2026),
# y hydro_reservoir_mw sola tiene un escalon de nivel en esa misma frontera (media 2.756 en 2021,
# 1.720 en 2022, 1.940 en 2023) -- exactamente el tipo de discontinuidad a mitad de train que ya
# encontramos antes con el autoconsumo, solo que aqui la fuente misma cambia de definicion.
TABLA_ENTSOE_REAL = "entsoe_gen_data"
COLS_ENTSOE_REAL = ["wind_mw", "pumping_cons_mw", "hydro_run_river_mw",
                     "biomass_mw", "waste_mw", "other_renewable_mw", "oil_mw"]

# ESIOS gana termosolar, CCGT puro, cogeneracion y resto (SIN dividir -- ver nota de
# _features_lag_reales), nuclear, carbon, y bateria (descarga/carga). ree_gnuclear_mw y
# ree_gcoal_mw van por ESIOS, no ENTSO-E, aunque coinciden al ±0,4% -- es la fuente que eligio el
# equipo (corregido 20-ago-2026: una version anterior de este patch los tenia por ENTSO-E).
TABLA_ESIOS_REAL = "esios_gen"
COLS_ESIOS_REAL = ["ree_gsolter_mw", "ree_gccgas_mw", "ree_gbattery_mw", "ree_cbattery_mw",
                    "ree_gnuclear_mw", "ree_gotherthermal_mw", "ree_gcoal_mw"]
# ree_gsolar_mw (FV) se EXCLUYE de COLS_ESIOS_REAL a proposito: incorpora autoconsumo desde
# dic-2025, mismo problema que ree_load (verificado: brecha de -546 a +1.387 MW entre sep-2025 y
# ago-2026). Se reconstruye una FV limpia por resta en vez de leerla directo -- ver
# _features_lag_reales: GREATEST(0, entsoe_gen_data.solar_mw - esios_gen.ree_gsolter_mw). ENTSO-E
# solar_mw es B16 = FV+termosolar sin el problema de autoconsumo (es de ENTSO-E, no de ESIOS);
# restarle la termosolar limpia de ESIOS deja una FV limpia.
#
# ree_gotherthermal_mw (cogeneracion y resto, indicador 1297) entra COMPLETA, sin dividir en
# cogeneracion_gas_mw/resto_no_convencional_mw como se planteo en una version anterior
# (gas_lags.patch): el equipo evaluo esa division a fondo y la descarto -- la resta sigue las
# rampas del CCGT (57 MW a las 9h, 332 MW a las 19h) porque en realidad mide el desfase entre una
# CONSIGNA en tiempo real (entsoe.gas_mw, B04) y una TELEMEDIDA real (ree_gccgas_mw), no una
# tecnologia; resto_no_convencional_mw llegaba a ser negativa 10 veces su propia magnitud en el
# 11% de las horas de 2025. Ver docs/decisiones_datos.md D-02 y la matriz FINAL del 19-ago.


def _ruta_versionada(carpeta: Path, nombre: str, ext: str = ".csv",
                     sufijo: str = "", sobrescribir: bool = False) -> Path:
    """Devuelve una ruta libre, versionando en vez de pisar lo que ya hay.

    Añadido 23-ago-2026. Con varias personas generando el dataset y el catalogo de columnas
    todavia moviendose, sobrescribir el CSV anterior significa perder la referencia contra la
    que alguien ya habia entrenado -- y no poder explicar de donde salio un resultado.

    Numera SIEMPRE, desde _v01, para que el orden alfabetico coincida con el cronologico:
        dataset_horario_v01.csv, dataset_horario_v02.csv, ...
    Con `sufijo` (opcional, vacio por defecto) se separan lineas de trabajo dentro de una misma
    carpeta. Normalmente no hace falta: el nombre ya lleva la variante y la carpeta identifica
    la linea de trabajo.
    Con `sobrescribir=True` se reutiliza la ultima version existente (util al iterar sobre un
    bug, para no dejar veinte ficheros identicos).
    """
    carpeta.mkdir(parents=True, exist_ok=True)
    tronco = f"{nombre}_{sufijo}" if sufijo else nombre
    existentes = sorted(carpeta.glob(f"{tronco}_v[0-9][0-9]{ext}"))
    if sobrescribir and existentes:
        return existentes[-1]
    n = 1
    if existentes:
        try:
            n = int(existentes[-1].stem.rsplit("_v", 1)[1]) + 1
        except (IndexError, ValueError):
            n = len(existentes) + 1
    if n > 99:
        raise RuntimeError(f"99 versiones de {tronco} en {carpeta}; limpia antes de seguir")
    return carpeta / f"{tronco}_v{n:02d}{ext}"


def _escribir_con_meta(df: pd.DataFrame, ruta: Path, descripcion: str,
                       float_format: str = "%.4f", index: bool = False,
                       index_label: str | None = None, **extra) -> Path:
    """Escribe el CSV y, al lado, un .meta.json con de donde sale.

    Sin esto el versionado solo produce confusion: cinco ficheros con nombres casi iguales y
    nadie sabe cual llevaba PDBC o cual se genero con el clima activado. El sidecar guarda las
    constantes que definen el dataset y la lista completa de columnas.
    """
    import json
    from datetime import datetime as _dt

    df.to_csv(ruta, index=index, index_label=index_label, float_format=float_format)
    meta = {
        "descripcion": descripcion,
        "generado": _dt.now().isoformat(timespec="seconds"),
        "fichero": ruta.name,
        "filas": int(df.shape[0]),
        "columnas": int(df.shape[1]),
        "DATASET_START": DATASET_START,
        "DATASET_END": DATASET_END,
        "MODELO_END": str(MODELO_END),
        "TRAIN_END": str(TRAIN_END),
        "VAL_END": str(VAL_END),
        "TOPE_GAS": [str(TOPE_GAS_INI.date()), str(TOPE_GAS_FIN.date())],
        "capacidad_como_cuota": CAPACIDAD_COMO_CUOTA,
        "filtrar_tendencias": FILTRAR_TENDENCIAS,
        "umbral_tendencia": UMBRAL_TENDENCIA,
        "cobertura_minima_train": COBERTURA_MINIMA_TRAIN,
        "periodos_excluidos": PERIODOS_EXCLUIDOS,
    }
    meta.update(extra)
    meta["lista_columnas"] = list(df.columns)
    ruta.with_suffix(".meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
    return ruta


def _conectar():
    _, db_config = load_config()
    return psycopg2.connect(**db_config)


def _columnas_de(conn, tabla: str) -> list:
    """Nombres reales de columna de una tabla/vista, en orden.

    Añadido 23-ago-2026. Motivo: la matriz FINAL del equipo (Excel) y la BD no siempre usan el
    mismo nombre para la misma columna -- el Excel lista `ree_demanda_prev` donde el script lee
    `demanda_mercado_prev_mw`, por ejemplo. Al incorporar los bloques que faltaban (forecast de
    autoconsumo, PDBC, capacidad) preferimos filtrar contra lo que EXISTE de verdad y avisar de
    lo que no aparece, en vez de que el SELECT reviente con un UndefinedColumn a mitad de carga.
    """
    df = pd.read_sql(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_name = %(t)s ORDER BY ordinal_position",
        conn, params={"t": tabla},
    )
    return df["column_name"].tolist()


def _filtrar_existentes(conn, tabla: str, deseadas: list, etiqueta: str) -> list:
    """Interseccion entre lo que se quiere y lo que hay, avisando de la diferencia."""
    reales = set(_columnas_de(conn, tabla))
    ok = [c for c in deseadas if c in reales]
    falta = [c for c in deseadas if c not in reales]
    if falta:
        print(f"AVISO [{etiqueta}]: {len(falta)} columnas no existen en `{tabla}` y se omiten: {falta}")
    return ok


def _col_temporal(conn, tabla: str) -> str:
    """Devuelve la columna temporal de la tabla. La BD todavia no esta normalizada: conviven
    `datetime`, `date`, `fecha`, `ts` y `time_qh` segun la tabla (revision global 17-ago-2026)."""
    reales = _columnas_de(conn, tabla)
    for cand in ("datetime", "date", "fecha", "ts", "time_qh"):
        if cand in reales:
            return cand
    raise ValueError(f"No encuentro columna temporal en `{tabla}`: {reales[:10]}")


def _a_fecha_local(serie: pd.Series) -> pd.Series:
    """Columna temporal -> fecha civil de Madrid, funcione la columna como sea.

    Añadido 23-ago-2026 tras un crash real en `esios_capacity_available`:
        ValueError: Tz-aware datetime.datetime cannot be converted to datetime64
                    unless utc=True, at position 2034
    La posicion 2034 era un cambio de hora. psycopg2 devuelve datetime tz-aware con offset
    +01:00 o +02:00 segun la epoca del año, pandas recibe un array de objetos con DOS offsets
    distintos y no puede inferir un dtype unico. Es el mismo fenomeno que documenta
    `_leer_horaria` para parse_dates.

    Y NO vale poner utc=True a secas: si la columna es un `date` naive (o un timestamp sin
    huso), interpretarlo como UTC y pasarlo a Madrid resta una hora y la fecha retrocede un dia
    -- que es exactamente el desplazamiento de toda la serie historica que se corrigio en esta
    misma tabla el 17-ago-2026. Por eso se decide segun el valor real:

        aware  -> to_datetime(utc=True) -> tz_convert(Madrid) -> .date
        naive  -> to_datetime()                               -> .date   (se usa tal cual)
    """
    limpia = serie.dropna()
    if limpia.empty:
        return pd.to_datetime(serie).dt.date
    aware = getattr(limpia.iloc[0], "tzinfo", None) is not None
    if aware:
        return pd.to_datetime(serie, utc=True).dt.tz_convert("Europe/Madrid").dt.date
    return pd.to_datetime(serie).dt.date


def _leer_horaria(conn, sql: str, col_ts: str, params: dict) -> pd.DataFrame:
    """SELECT + normalizacion UTC manual. NUNCA usar parse_dates=[...] aqui -- sobre una columna
    timestamptz que cruza un cambio de hora (CET/CEST), parse_dates la deja en un offset FIJO y
    convierte en NaT la mitad de las filas (las del otro offset). Un merge posterior por esa
    columna hace que esos NaT casen entre si y el join explota (visto en este proyecto: un join
    de 31.583 x 31.583 filas intento reservar 48 GB). Por eso se lee sin parse_dates y se
    normaliza a mano con pd.to_datetime(..., utc=True) justo despues."""
    df = pd.read_sql(sql, conn, params=params)
    df[col_ts] = pd.to_datetime(df[col_ts], utc=True)
    return df


def construir_espina_horaria(start: str = DATASET_START, end: str = DATASET_END) -> pd.DataFrame:
    """Espina horaria de EXPLORACION (correlaciones, distribuciones, graficos) -- no es el
    dataset de modelado (para eso, `construir_dataset_diario`). Union por timestamp exacto de
    las 5 tablas nucleo (ver Banco de Evidencias / notas de negocio para el criterio de por que
    estas 5), con `entsoe_gen_data` como eje porque es la mas completa/limpia. Se añaden ademas
    las variables diarias (commodities, capacidad) difundidas sobre las 24h del dia.
    """
    conn = _conectar()
    try:
        params = {"start": start, "end": end}
        df_gen = _leer_horaria(
            conn, "SELECT * FROM entsoe_gen_data WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            "datetime", params,
        )
        df_load = _leer_horaria(
            conn, "SELECT * FROM entsoe_load_inter WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            "datetime", params,
        )
        df_esios_gen = _leer_horaria(
            conn, "SELECT * FROM esios_gen WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            "datetime", params,
        )
        df_load_inter = _leer_horaria(
            conn, "SELECT * FROM load_inter WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            "datetime", params,
        )
        df_forecast = _leer_horaria(
            conn, "SELECT * FROM esios_forecast_da WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            "datetime", params,
        )
        df_commodities = pd.read_sql(
            "SELECT * FROM commodities WHERE fecha BETWEEN %(start)s AND %(end)s ORDER BY fecha", conn, params=params,
        ).sort_values("fecha").ffill()   # fines de semana/festivos -> ultimo cierre, no interpolar
        df_commodities["fecha"] = pd.to_datetime(df_commodities["fecha"]).dt.date
        df_cap_disp = pd.read_sql(
            "SELECT * FROM esios_capacity_available WHERE date BETWEEN %(start)s AND %(end)s ORDER BY date",
            conn, params=params,
        )
        df_cap_inst = pd.read_sql(
            "SELECT * FROM esios_capacity_installed WHERE date BETWEEN %(start)s AND %(end)s ORDER BY date",
            conn, params=params,
        )
    finally:
        conn.close()

    espina = df_gen.rename(columns={"datetime": "ts"})
    espina = espina.merge(df_load.rename(columns={"datetime": "ts"}), on="ts", how="left", suffixes=("", "_load"))
    espina = espina.merge(df_esios_gen.rename(columns={"datetime": "ts"}), on="ts", how="left", suffixes=("", "_esiosgen"))
    espina = espina.merge(df_load_inter.rename(columns={"datetime": "ts"}), on="ts", how="left", suffixes=("", "_loadinter"))
    espina = espina.merge(df_forecast.rename(columns={"datetime": "ts"}), on="ts", how="left", suffixes=("", "_fcst"))
    assert espina["ts"].is_unique, "la espina no deberia tener timestamps duplicados -- revisar duplicados de fuente"

    espina["fecha"] = espina["ts"].dt.tz_convert("UTC").dt.date
    for nombre, df_diario, col_fecha in [
        ("commodities", df_commodities, "fecha"),
        ("esios_capacity_available", df_cap_disp, "date"),
        ("esios_capacity_installed", df_cap_inst, "date"),
    ]:
        df_diario = df_diario.copy()
        df_diario["fecha"] = pd.to_datetime(df_diario[col_fecha]).dt.date
        if col_fecha != "fecha":
            df_diario = df_diario.drop(columns=[col_fecha])
        espina = espina.merge(df_diario, on="fecha", how="left", suffixes=("", f"_{nombre}"))

    assert espina["ts"].is_unique, "la espina no deberia tener timestamps duplicados tras el join diario"
    return espina


def _construir_lag(df: pd.DataFrame, dias: int, sufijo: str) -> pd.DataFrame:
    """Desplaza el indice HACIA ADELANTE `dias` dias: el dato descrito en la fecha X
    aparece en la fila X + dias (para que quede alineado a la fila que lo usa como lag)."""
    out = df.copy()
    out.index = pd.to_datetime(out.index) + pd.Timedelta(days=dias)
    out.index = out.index.date
    out.columns = [f"{c}_{sufijo}" for c in out.columns]
    out.index.name = "fecha"
    return out


def _target_d1(conn) -> pd.DataFrame:
    """Precio real horario -> 24 columnas por dia, YA DESPLAZADO para que la fila D
    contenga el precio real de D+1 (no el de D -- ver nota de correccion del 17-ago-2026)."""
    df = pd.read_sql(
        "SELECT datetime, es_esios::float8 AS es_esios FROM spot_price "
        "WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)  # normalizacion UTC manual, NUNCA parse_dates
    df["fecha_madrid"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date
    df["hora"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.hour

    wide = df.pivot_table(index="fecha_madrid", columns="hora", values="es_esios", aggfunc="first")
    wide.columns = [f"price_h{h:02d}" for h in wide.columns]
    wide.index.name = "fecha"

    idx_completo = pd.date_range(pd.to_datetime(wide.index.min()), pd.to_datetime(wide.index.max()), freq="D")
    wide.index = pd.to_datetime(wide.index)
    target_d1 = wide.reindex(idx_completo).shift(-1)   # fila D <- precio real de D+1
    target_d1.index = target_d1.index.date
    target_d1.index.name = "fecha"
    return target_d1


def _features_forecast(conn) -> pd.DataFrame:
    """Previsiones oficiales de esios_forecast_da (seguras: publicadas antes del cierre del
    mercado), agregadas por dia y alineadas a la fila D que las conoce."""
    reales = set(_columnas_de(conn, "esios_forecast_da"))
    cols = _filtrar_existentes(conn, "esios_forecast_da",
                               COLS_SEGURAS_FORECAST + COLS_SEGURAS_FORECAST_EXTRA, "forecast")
    # autoconsumo previsto: se coge el primer alias que exista
    extra = [c for c in COLS_FORECAST_AUTOCONSUMO if c in reales][:1]
    cols = cols + extra
    select = ", ".join(cols)
    if COL_FORECAST_FLAG and COL_FORECAST_FLAG in reales:
        select += f", {COL_FORECAST_FLAG}::int AS {COL_FORECAST_FLAG}"
        cols = cols + [COL_FORECAST_FLAG]

    df = pd.read_sql(
        f"SELECT datetime, {select} FROM esios_forecast_da "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha_objetivo"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date  # dia D+1 que predice

    # c_autoconsumo_prev reconstruida (ruta DIARIA). Respeta EXCLUIR_AUTOCONSUMO: si no, el
    # dataset diario y el horario llevarian columnas distintas y la comparacion entre ambos
    # dejaria de ser valida.
    if not EXCLUIR_AUTOCONSUMO and {"demanda_prev_mw", "demanda_mercado_prev_mw"} <= set(df.columns):
        df["c_autoconsumo_prev"] = df["demanda_prev_mw"] - df["demanda_mercado_prev_mw"]
        cols = cols + ["c_autoconsumo_prev"]
    elif not EXCLUIR_AUTOCONSUMO:
        print("AVISO [forecast]: faltan 1775 y/o 2563, no se puede reconstruir c_autoconsumo_prev")

    feats = df.groupby("fecha_objetivo")[cols].agg(["mean", "min", "max"])
    feats.columns = [f"{c}_{stat}" for c, stat in feats.columns]
    feats.index = (pd.to_datetime(feats.index) - pd.Timedelta(days=1)).date  # -> fila del dia D
    feats.index.name = "fecha"
    return feats


# NTC (capacidad de interconexion) -- ganador del punto #3, publicada antes del cierre igual que
# las previsiones. Ahora en `load_inter` (antes esios_load_inter, eliminada 19-ago-2026). Marruecos
# (ree_ntc_impma/expma) se incluye desde el 20-ago-2026 a peticion del equipo -- antes se dejaba
# fuera porque no se habia resuelto en el punto #3. NOTA: es capacidad publicada, no flujo real
# (eso va en COLS_DEMANDA_REAL, ver arriba) -- por eso se alinea D+1 como las previsiones.
COLS_NTC = ["ree_ntc_impfr", "ree_ntc_expfr", "ree_ntc_imppt", "ree_ntc_exppt",
            "ree_ntc_impma", "ree_ntc_expma"]

# entsoe_forecast_da: revivida el 17-ago-2026 (antes 192 filas, ahora historico completo).
# renewables_forecast_mw se EXCLUYE por el mismo motivo que gen_renovables_prev_mw: es
# wind_forecast_mw + solar_forecast_mw exacto (verificado 18-ago-2026, 57.598 filas identicas).
COLS_FORECAST_ENTSOE = ["load_forecast_mw", "wind_forecast_mw", "solar_forecast_mw"]


def _features_ntc(conn) -> pd.DataFrame:
    """NTC Francia/Portugal de load_inter -- mismo criterio de alineacion que
    _features_forecast (publicada antes del cierre, describe D+1, se coloca en la fila D)."""
    df = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_NTC)} FROM load_inter WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha_objetivo"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date

    feats = df.groupby("fecha_objetivo")[COLS_NTC].agg(["mean", "min", "max"])
    feats.columns = [f"{c}_{stat}" for c, stat in feats.columns]
    feats.index = (pd.to_datetime(feats.index) - pd.Timedelta(days=1)).date
    feats.index.name = "fecha"
    return feats


def _features_forecast_entsoe(conn) -> pd.DataFrame:
    """Segunda prevision oficial (ENTSO-E, independiente de ESIOS) -- misma alineacion D+1.
    FUERA del dataset por defecto desde 20-ago-2026 -- decision de reunion del equipo, ver
    docs/columnas_pendientes_equipo.md. Solo se activa con `incluir_columnas_pendientes=True`."""
    df = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_FORECAST_ENTSOE)} FROM entsoe_forecast_da "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha_objetivo"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date

    feats = df.groupby("fecha_objetivo")[COLS_FORECAST_ENTSOE].agg(["mean", "min", "max"])
    feats.columns = [f"entsoe_{c}_{stat}" for c, stat in feats.columns]
    feats.index = (pd.to_datetime(feats.index) - pd.Timedelta(days=1)).date
    feats.index.name = "fecha"
    return feats


def _features_diferencia_previsiones(conn) -> pd.DataFrame:
    """Diferencia horaria entre las DOS previsiones independientes (ESIOS vs ENTSO-E) para
    demanda/eolica/solar, agregada por dia -- proxy de incertidumbre del dia. Las dos se
    publican antes del cierre, asi que la diferencia en si tambien es segura como feature
    (no hay fuga: ninguna de las dos "sabe" mas que la otra sobre el resultado real).
    FUERA del dataset por defecto desde 20-ago-2026 -- decision de reunion del equipo, ver
    docs/columnas_pendientes_equipo.md. Depende de _features_forecast_entsoe, mismo interruptor."""
    df_esios = pd.read_sql(
        "SELECT datetime, demanda_prev_mw, gen_wind_prev_mw, gen_solar_pv_prev_mw FROM esios_forecast_da "
        "WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df_entsoe = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_FORECAST_ENTSOE)} FROM entsoe_forecast_da "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df_esios["datetime"] = pd.to_datetime(df_esios["datetime"], utc=True)
    df_entsoe["datetime"] = pd.to_datetime(df_entsoe["datetime"], utc=True)

    df = df_esios.merge(df_entsoe, on="datetime", how="inner")
    df["diff_demanda"] = (df["demanda_prev_mw"] - df["load_forecast_mw"]).abs()
    df["diff_eolica"] = (df["gen_wind_prev_mw"] - df["wind_forecast_mw"]).abs()
    df["diff_solar"] = (df["gen_solar_pv_prev_mw"] - df["solar_forecast_mw"]).abs()
    df["fecha_objetivo"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date

    cols_diff = ["diff_demanda", "diff_eolica", "diff_solar"]
    feats = df.groupby("fecha_objetivo")[cols_diff].agg(["mean", "max"])
    feats.columns = [f"{c}_{stat}" for c, stat in feats.columns]
    feats.index = (pd.to_datetime(feats.index) - pd.Timedelta(days=1)).date
    feats.index.name = "fecha"
    return feats


# 23-ago-2026: reducido a las 3 ACTIVAS de la matriz FINAL. gas_ttf, co2_ets y carbon_api2
# estan marcadas DESCARTADAS en el Excel del equipo (carbon_api2 ademas quedo vacia al
# deslistarse MTF=F, y co2_ets fue sustituida por co2_eua_dec). El script y la matriz decian
# cosas distintas; manda la matriz.
COLS_COMMODITIES = ["gas_mibgas", "co2_eua_dec", "gas_ttf_m1"]


def _features_dia_d(conn) -> pd.DataFrame:
    """Commodities -- CIERRE DE D-1 colocado en la fila D (corregido 23-ago-2026).

    Antes se usaba el cierre del propio dia D con el comentario "lag natural: ya ocurrieron".
    No es asi: la prediccion para D+1 se emite a las 12:00 de D, y el cierre de TTF/EUA/MIBGAS
    de D se produce hacia las 17:30 CET, es decir DESPUES del cierre del mercado electrico.
    El ultimo cierre realmente disponible al predecir es el de D-1. Fuga de ~5,5 h, pequeña en
    magnitud (el gas es muy persistente) pero real y trivial de corregir.

    Dos cambios mas:
      - ffill con limite de 4 dias (cubre puente largo). Sin limite, una caida de fuente se
        propaga como precio constante indefinidamente -- ya paso con un hueco de 13 dias en
        gas_ttf que no se detecto.
      - columna `dias_desde_cierre`: cuantos dias hace del ultimo dato real. Sirve de feature
        de frescura y de alarma en el EDA.
    """
    df_comm = pd.read_sql(
        f"SELECT fecha, {', '.join(COLS_COMMODITIES)} FROM commodities WHERE fecha BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df_comm["fecha"] = pd.to_datetime(df_comm["fecha"]).dt.date
    df_comm = df_comm.sort_values("fecha").set_index("fecha")

    # reindex a calendario completo ANTES del ffill, si no los findes no existen como filas
    idx = pd.date_range(pd.to_datetime(min(df_comm.index)), pd.to_datetime(max(df_comm.index)), freq="D")
    df_comm.index = pd.to_datetime(df_comm.index)
    df_comm = df_comm.reindex(idx)

    hay_dato = df_comm[COLS_COMMODITIES].notna().any(axis=1)
    df_comm = df_comm.ffill(limit=4)
    # dias transcurridos desde la ultima fila con dato real
    ultima = pd.Series(np.where(hay_dato, np.arange(len(df_comm)), np.nan), index=df_comm.index).ffill()
    df_comm["dias_desde_cierre"] = np.arange(len(df_comm)) - ultima

    # cierre de D-1 -> fila D
    df_comm.index = (df_comm.index + pd.Timedelta(days=1)).date
    df_comm.index.name = "fecha"

    fuera = [c for c in COLS_EXCLUIDAS if c in df_comm.columns]
    if fuera:
        df_comm = df_comm.drop(columns=fuera)
    return df_comm


def _features_capacidad_disponible(conn) -> pd.DataFrame:
    """Capacidad disponible del propio dia D (lag natural). FUERA del dataset por defecto desde
    20-ago-2026 -- decision de reunion del equipo, ver docs/columnas_pendientes_equipo.md. Solo
    se activa con `incluir_columnas_pendientes=True` en `construir_dataset_diario`."""
    df_capd = pd.read_sql(
        "SELECT date, total_mw FROM esios_capacity_available WHERE date BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df_capd["date"] = pd.to_datetime(df_capd["date"]).dt.date
    return df_capd.rename(columns={"total_mw": "capacidad_disp_total_mw"}).set_index("date")


# Excepcion iberica ("tope al gas", RDL 10/2022): mecanismo de ajuste que deprimio
# artificialmente el precio marginal español y portugues. Fechas a VERIFICAR contra BOE antes
# de la memoria. Son ~18 meses = 27% del historico, y caen ENTEROS dentro de train mientras
# validation (2025) y test (2026) quedan fuera: sin este dummy el modelo aprende una relacion
# gas->precio promediada entre dos regimenes que no coexisten. Añadido 23-ago-2026.
TOPE_GAS_INI = pd.Timestamp("2022-06-15")
TOPE_GAS_FIN = pd.Timestamp("2023-12-31")

# Cambio metodologico de REE: incorporacion de la estimacion de autoconsumo (documento de
# metodologia de diciembre de 2025, seccion 3.2). Afecta en cascada a ree_load (1293), al
# indicador 1775 y a la solar FV de ESIOS (1295). Sustituye a la columna `autoconsumo_estimado`
# de la matriz, que NO existe en la BD: es un regimen que solo depende de la fecha, asi que un
# dummy de calendario lo representa igual de bien y sin depender de la ingesta.
# Cae dentro de validation/test pero NO de train -- por eso importa marcarlo.
# FECHA A AFINAR: la brecha ree_load - entsoe_load pasa de +-1 MW a ~435 MW en ese salto.
#   SELECT date_trunc('day', datetime) AS dia, avg(ree_load - entsoe_load) AS brecha_mw
#   FROM load_inter WHERE datetime >= '2025-11-15' AND datetime < '2026-01-15'
#   GROUP BY 1 ORDER BY 1;
# El primer dia con brecha de centenares de MW es el valor que debe ir aqui.
AUTOCONSUMO_INI = pd.Timestamp("2025-12-01")   # verificado en BD, corte limpio


# ══ FESTIVOS NACIONALES DE ESPAÑA (23-ago-2026) ══════════════════════════════════════════
# El paquete `holidays` nunca llego a instalarse y d1_is_festivo salia constante a 0 en todas
# las ejecuciones, perdiendo una feature con peso real: en precio electrico el 6 de diciembre o
# el 15 de agosto se comportan como un domingo, y `d1_is_weekend` no los ve.
#
# Se incluye el calendario laboral nacional calculado, sin dependencias externas. Las moviles
# (Jueves y Viernes Santo) se derivan de la Pascua con el algoritmo gregoriano anonimo
# (Meeus/Butcher), asi que vale para cualquier año sin mantenimiento.
#
# ALCANCE: solo festivos NACIONALES. Los autonomicos y locales quedan fuera a proposito -- el
# target es el precio del sistema peninsular, y un festivo de una sola comunidad apenas mueve
# la demanda agregada. El Jueves Santo si entra aunque no sea festivo en Cataluña ni en la
# Comunidad Valenciana: en el resto del pais lo es, y el efecto sobre la demanda peninsular es
# claro. El Lunes de Pascua (Cataluña, Valencia, Baleares, Navarra, La Rioja, Pais Vasco) se
# deja fuera por ser mayoritariamente autonomico.
#
# Nota metodologica: se usan las fechas REALES, sin aplicar los traslados del calendario
# laboral cuando un festivo cae en domingo. Lo que mueve la demanda es el dia que la gente no
# trabaja, y ese traslado varia por comunidad.

def _pascua(anio: int):
    """Domingo de Pascua (algoritmo gregoriano anonimo, Meeus/Butcher)."""
    a = anio % 19
    b, c = divmod(anio, 100)
    d, e = divmod(b, 4)
    g = (8 * b + 13) // 25
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes, dia = divmod(h + l - 7 * m + 114, 31)
    return pd.Timestamp(anio, mes, dia + 1)


def festivos_nacionales(anios) -> set:
    """Conjunto de `date` con los festivos nacionales de España de los años indicados."""
    fijos = [(1, 1),    # Año Nuevo
             (1, 6),    # Epifania del Señor
             (5, 1),    # Fiesta del Trabajo
             (8, 15),   # Asuncion de la Virgen
             (10, 12),  # Fiesta Nacional de España
             (11, 1),   # Todos los Santos
             (12, 6),   # Dia de la Constitucion
             (12, 8),   # Inmaculada Concepcion
             (12, 25)]  # Natividad del Señor
    out = set()
    for y in anios:
        for mes, dia in fijos:
            out.add(pd.Timestamp(y, mes, dia).date())
        p = _pascua(y)
        out.add((p - pd.Timedelta(days=3)).date())   # Jueves Santo
        out.add((p - pd.Timedelta(days=2)).date())   # Viernes Santo
    return out


# ══ PDBC ═════════════════════════════════════════════════════════════════════════════════
# Bloque de la matriz FINAL que faltaba por completo en el script (23-ago-2026).
# CRITICO: el PDBC del dia D+1 NO se puede usar. Se publica hacia las 13:00 de D, posterior al
# cierre de las 12:00, y ademas sale de la MISMA casacion que el precio -- usarlo seria tan
# circular como usar el target. Por eso entra unicamente con lag D-1 y D-7, igual que el resto
# de series reales. Su valor real es de interpretabilidad (desviacion programa vs. real), no de
# prediccion; conviene tenerlo presente al leer las importancias de variables.
COLS_PDBC = ["wind_mw", "solar_pv_mw", "solar_thermal_mw", "hydro_ugh_mw", "hydro_no_ugh_mw",
             "nuclear_mw", "coal_mw", "cogen_mw", "biomass_mw", "biogas_mw", "hybrid_mw",
             "ccgt_mw", "fuel_gas_mw", "waste_mw", "other_renew_mw", "pumping_gen_mw",
             "pumping_cons_mw"]

# ══ CAPACIDAD ════════════════════════════════════════════════════════════════════════════
# Las dos tablas de capacidad estaban fuera del dataset diario (la disponible solo aparecia
# como `total_mw` y detras del flag apagado; la instalada ni eso). Se incorporan enteras a
# peticion del equipo, 23-ago-2026.
#
# ALINEACION: las dos son magnitudes PUBLICADAS POR ADELANTADO, no dato realizado, asi que van
# a D+1 igual que las previsiones y las NTC -- no con lag. Para la disponible esta verificado
# que el valor D+1 no tiene variacion intradiaria (24 horas identicas); la version revisada a
# D-2 si varia hasta 838 MW, pero esa no es la que se usa al predecir.
#
# SALVEDAD sobre la instalada: 23 columnas de una serie que se mueve por escalones mensuales y
# fuertemente colineales entre si (total_renewable_mw es la suma de varias de las demas). Se
# cargan todas porque asi lo pide la matriz, pero al hacer seleccion de variables lo esperable
# es que sobrevivan 4 (total_renewable_mw, solar_pv_mw, wind_mw, battery_hybrid_mw), que son las
# que capturan la tendencia de penetracion renovable. El resto es redundancia correlacionada.
COLS_CAP_INSTALADA = ["total_mw", "total_renewable_mw", "total_nonrenewable_mw", "total_hybrid_mw",
                      "total_autoconsume_mw", "hydro_mw", "pump_mw", "wind_mw", "wind_hybrid_mw",
                      "solar_pv_mw", "solar_thermal_mw", "solar_pv_hybrid_mw", "other_renewable_mw",
                      "waste_nonrenewable_mw", "waste_renewable_mw", "battery_hybrid_mw",
                      "autoconsume_solar_pv_mw", "autoconsume_battery_mw", "nuclear_mw",
                      "coal_mw", "fuel_mw", "ccgt_mw", "cogeneration_mw"]

# `total_mw` NO existe en esta tabla (verificado 23-ago-2026 contra la BD); son las 6 de la
# matriz y ninguna mas.
COLS_CAP_DISPONIBLE = ["hydro_mw", "pump_mw", "nuclear_mw", "coal_antracita_mw",
                       "ccgt_mw", "fuel_mw"]


def _features_pdbc(conn) -> pd.DataFrame:
    """PDBC por tecnologia, agregado por dia y desplazado a D-1 y D-7. Nunca D ni D+1."""
    cols = _filtrar_existentes(conn, "esios_pdbc_gen", COLS_PDBC, "pdbc")
    if not cols:
        print("AVISO [pdbc]: `esios_pdbc_gen` sin columnas utiles, bloque omitido")
        return pd.DataFrame()
    df = pd.read_sql(
        f"SELECT datetime, {', '.join(cols)} FROM esios_pdbc_gen "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date

    stats = PDBC_ESTADISTICOS if PODAR_REDUNDANTES else ["mean", "min", "max"]
    if PODAR_REDUNDANTES:
        cols = [c for c in cols if any(t in c for t in PDBC_TECNOLOGIAS_UTILES)]

    diario = df.groupby("fecha")[cols].agg(stats)
    diario.columns = [f"pdbc_{c}_{stat}" for c, stat in diario.columns]

    lags = PDBC_LAGS if PODAR_REDUNDANTES else ["lag1d", "lag7d"]
    piezas = [_construir_lag(diario, {"lag1d": 1, "lag7d": 7}[l], l) for l in lags]
    out = piezas[0] if len(piezas) == 1 else piezas[0].join(piezas[1:], how="outer")
    print(f"[pdbc] {out.shape[1]} columnas ({len(cols)} tecnologias x {len(stats)} stats "
          f"x {len(lags)} lags)")
    return out


def _features_capacidad(conn, tabla: str, cols_deseadas: list, prefijo: str) -> pd.DataFrame:
    """Capacidad (instalada o disponible) del dia D+1, colocada en la fila D.

    Una fila por dia en origen, asi que no hay agregacion horaria: se renombra y se desplaza.
    La columna temporal se detecta en runtime porque la BD aun no esta normalizada (`date` en
    una tabla, `datetime` en la otra segun la matriz).
    """
    cols = _filtrar_existentes(conn, tabla, cols_deseadas, prefijo)
    if not cols:
        print(f"AVISO [{prefijo}]: `{tabla}` sin columnas utiles, bloque omitido")
        return pd.DataFrame()
    ts = _col_temporal(conn, tabla)
    df = pd.read_sql(
        f"SELECT {ts}, {', '.join(cols)} FROM {tabla} WHERE {ts} BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["fecha_objetivo"] = _a_fecha_local(df[ts])

    out = df.groupby("fecha_objetivo")[cols].mean()   # 1 fila/dia; mean() solo colapsa duplicados
    out.columns = [f"{prefijo}_{c}" for c in out.columns]
    out.index = (pd.to_datetime(out.index) - pd.Timedelta(days=1)).date   # D+1 -> fila D
    out.index.name = "fecha"
    return out


def _calendario(index_fechas) -> pd.DataFrame:
    """Calendario de D+1 (determinista, siempre conocido de antemano).

    Ampliado 23-ago-2026 con tres bloques:
      - FESTIVOS nacionales. `d1_is_weekend` no captura el 6-dic, el 15-ago ni Semana Santa,
        que en precio electrico pesan tanto o mas que un sabado. Requiere `pip install holidays`
        (paquete puro Python, sin dependencias); si no esta, la columna sale a 0 y se avisa.
      - CODIFICACION CICLICA (seno/coseno) de dia de semana y mes. Para arboles da igual, pero
        una red neuronal con `d1_month` entero aprende que diciembre esta lejos de enero.
        Las columnas enteras se mantienen para no romper a quien ya las use.
      - DUMMY DE REGIMEN del tope al gas (ver constantes arriba).
    """
    cal = pd.DataFrame(index=index_fechas)
    d1 = pd.to_datetime(cal.index) + pd.Timedelta(days=1)
    cal["d1_dow"] = d1.dayofweek
    cal["d1_month"] = d1.month
    cal["d1_is_weekend"] = (d1.dayofweek >= 5).astype(int)

    cal["d1_dow_sin"] = np.sin(2 * np.pi * d1.dayofweek / 7)
    cal["d1_dow_cos"] = np.cos(2 * np.pi * d1.dayofweek / 7)
    cal["d1_month_sin"] = np.sin(2 * np.pi * (d1.month - 1) / 12)
    cal["d1_month_cos"] = np.cos(2 * np.pi * (d1.month - 1) / 12)
    cal["d1_doy_sin"] = np.sin(2 * np.pi * d1.dayofyear / 365.25)
    cal["d1_doy_cos"] = np.cos(2 * np.pi * d1.dayofyear / 365.25)

    fest = festivos_nacionales(range(int(d1.year.min()), int(d1.year.max()) + 1))
    es_fest = np.array([d.date() in fest for d in d1])

    cal["d1_is_festivo"] = es_fest.astype(int)
    # puente: laborable atrapado entre festivo/finde por los dos lados
    no_lab = es_fest | (d1.dayofweek >= 5)
    prev = np.r_[False, no_lab[:-1]]
    sig = np.r_[no_lab[1:], False]
    cal["d1_es_puente"] = (~no_lab & prev & sig).astype(int)
    # vispera de festivo: la tarde/noche previa a un festivo tiene perfil propio (menos
    # industria, mas residencial), y no la capturan ni is_festivo ni is_weekend
    cal["d1_vispera_festivo"] = np.r_[es_fest[1:], False].astype(int)

    cal["d1_regimen_tope_gas"] = ((d1 >= TOPE_GAS_INI) & (d1 <= TOPE_GAS_FIN)).astype(int)
    # d1_regimen_autoconsumo se omite por el mismo motivo que autoconsumo_estimado: el cambio
    # es de dic-2025 y train acaba en dic-2024, asi que en train es constante. Se conserva la
    # constante AUTOCONSUMO_INI porque documenta la fecha exacta del corte para la memoria.
    if not EXCLUIR_AUTOCONSUMO:
        cal["d1_regimen_autoconsumo"] = (d1 >= AUTOCONSUMO_INI).astype(int)
    return cal


def _features_lag_reales(conn) -> pd.DataFrame:
    """Demanda, eolica/bombeo/hidraulica, termosolar+CCGT, y FV limpia (reconstruida), todas
    reales -- agregadas por dia y desplazadas a D-1 y D-7. Nunca el propio dia D (que aun no ha
    terminado al predecir)."""
    df_load = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_DEMANDA_REAL)} FROM {TABLA_DEMANDA_REAL} "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    # entsoe_load en 0 exacto es fisicamente imposible para demanda peninsular (verificado
    # 21-ago-2026: 9 horas de 58.151, 8 seguidas el 1-jul-2026 madrugada + 1 aislada el
    # 17-mar-2026 -- corte de ingesta puntual, no un problema sistemico. Caen en el tramo de
    # test, asi que sin este fix corromperian los lags D-1/D-7 de esos dias especificos con un
    # cero irreal. Se pasan a NaN para que la agregacion diaria (mean/min/max) las ignore, en vez
    # de tratarlas como demanda real de 0 MW).
    if "entsoe_load" in df_load.columns:
        df_load.loc[df_load["entsoe_load"] == 0, "entsoe_load"] = pd.NA
    df_entsoe = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_ENTSOE_REAL)} FROM {TABLA_ENTSOE_REAL} "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df_esios = pd.read_sql(
        f"SELECT datetime, {', '.join(COLS_ESIOS_REAL)} FROM {TABLA_ESIOS_REAL} "
        f"WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    # FV limpia: NUNCA se lee ree_gsolar_mw directo (contaminada de autoconsumo desde dic-2025).
    # Se reconstruye por resta: GREATEST(0, entsoe.solar_mw - esios.ree_gsolter_mw). El GREATEST
    # corrige los casos (~1% de las horas) donde la resta da un valor negativo por ruido de
    # redondeo entre las dos fuentes.
    df_solar_fv = pd.read_sql(
        "SELECT e.datetime, GREATEST(0, e.solar_mw - s.ree_gsolter_mw) AS solar_fv_mw "
        "FROM entsoe_gen_data e JOIN esios_gen s ON e.datetime = s.datetime "
        "WHERE e.datetime BETWEEN %(start)s AND %(end)s AND e.solar_mw IS NOT NULL AND s.ree_gsolter_mw IS NOT NULL",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    # Hidraulica despachable, fusionada (equivalente a c_ghydrodispatch de la matriz del equipo):
    # embalse (B12) + turbinacion de bombeo (B10), con COALESCE a 0 SOLO en el bombeo -- si el
    # embalse mismo es NULL, el resultado se queda NULL (no se inventa un cero). Necesario porque
    # ENTSO-E no separaba B10 de B12 antes de dic-2022 -- ver nota en COLS_ENTSOE_REAL.
    df_hydro_dispatch = pd.read_sql(
        "SELECT datetime, hydro_reservoir_mw + COALESCE(pumping_gen_mw, 0) AS hydro_dispatch_mw "
        "FROM entsoe_gen_data WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )

    piezas_horarias = [
        (df_load, COLS_DEMANDA_REAL), (df_entsoe, COLS_ENTSOE_REAL),
        (df_esios, COLS_ESIOS_REAL), (df_solar_fv, ["solar_fv_mw"]),
        (df_hydro_dispatch, ["hydro_dispatch_mw"]),
    ]
    aggs = []
    for df, cols in piezas_horarias:
        df["datetime"] = pd.to_datetime(df["datetime"], utc=True)  # normalizacion UTC manual, NUNCA parse_dates
        df["fecha"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date
        aggs.append(df.groupby("fecha")[cols].agg(["mean", "min", "max"]))

    real_diario = aggs[0].join(aggs[1:], how="outer")
    real_diario.columns = [f"{c}_{stat}" for c, stat in real_diario.columns]

    lag1 = _construir_lag(real_diario, 1, "lag1d")
    lag7 = _construir_lag(real_diario, 7, "lag7d")
    return lag1.join(lag7, how="outer")


def _features_lag_precio(target_wide_sin_shift: pd.DataFrame) -> pd.DataFrame:
    """Precio real como feature -- la señal de autocorrelacion mas fuerte del dataset.

    Reescrito 23-ago-2026. Dos cambios sobre la version anterior:

    1) SE AÑADE EL PRECIO DEL PROPIO DIA D (sufijo `_d0`, sin lag). Es legitimo: el precio de D
       se caso a las 12:00 de D-1 y se publico a las 13:00 de D-1, asi que a las 12:00 de D
       -- momento en que se predice D+1 -- lleva 23 h conocido. La version anterior solo usaba
       D-1 y D-7, descartando el dato autorregresivo mas fresco que existe. No es una fuga:
       es informacion que el operador tiene delante al pujar.

    2) SE AÑADE EL PERFIL HORARIO COMPLETO (24 columnas por dia), ademas de mean/min/max.
       Se predicen 24 valores; colapsar el precio de referencia a 3 estadisticos tira la forma
       de la curva, que es justo lo que hay que reproducir. price_h14 de D-7 explica price_h14
       de D+1 mucho mejor que la media del dia.

    Resultado: 6 columnas de estadisticos (D-1, D-7) + 72 de perfil horario (D, D-1, D-7).
    """
    precio_diario = target_wide_sin_shift.agg(["mean", "min", "max"], axis=1)
    precio_diario.columns = [f"precio_real_{stat}" for stat in precio_diario.columns]
    lag1 = _construir_lag(precio_diario, 1, "lag1d")
    lag7 = _construir_lag(precio_diario, 7, "lag7d")

    perfil_d0 = target_wide_sin_shift.copy()
    perfil_d0.columns = [f"{c}_d0" for c in perfil_d0.columns]
    perfil_d0.index.name = "fecha"
    perfil_lag1 = _construir_lag(target_wide_sin_shift, 1, "lag1d")
    perfil_lag7 = _construir_lag(target_wide_sin_shift, 7, "lag7d")

    return lag1.join([lag7, perfil_d0, perfil_lag1, perfil_lag7], how="outer")


# Precio de PT y FR del PROPIO dia D -- mismo razonamiento que el `_d0` de España: la subasta
# SDAC/MIBEL de D se resolvio a las 12:00 de D-1, asi que a las 12:00 de D ya se conoce.
# Lo que NO se puede usar es el de D+1 (se casa en la misma subasta simultanea que el target).
# `spread_es_pt_d0` es la feature interesante de las dos: PT correlaciona 0,997 con España y
# como nivel aporta poco, pero las horas en que los dos mercados se DESACOPLAN son las de
# congestion en la interconexion, que son las de precio extremo -- justo las que peor predice
# el modelo. El spread las señala; la columna cruda no.
def _features_precio_vecinos_d0(conn) -> pd.DataFrame:
    """Precio de PT y FR del dia D (conocido al predecir D+1) + spread ES-PT."""
    df = pd.read_sql(
        "SELECT datetime, es_esios::float8 AS es_esios, pt_entsoe::float8 AS pt_entsoe, "
        "fr_entsoe::float8 AS fr_entsoe FROM spot_price "
        "WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date
    df["spread_es_pt"] = df["es_esios"] - df["pt_entsoe"]
    df["spread_es_fr"] = df["es_esios"] - df["fr_entsoe"]

    cols = ["pt_entsoe", "fr_entsoe", "spread_es_pt", "spread_es_fr"]
    feats = df.groupby("fecha")[cols].agg(["mean", "min", "max"])
    feats.columns = [f"{c}_{stat}_d0" for c, stat in feats.columns]
    # horas de desacoplamiento con PT: proxy directo de congestion en la interconexion
    feats["horas_desacople_pt_d0"] = (
        df.assign(desac=(df["spread_es_pt"].abs() > 0.01)).groupby("fecha")["desac"].sum()
    )
    feats.index.name = "fecha"
    return feats


# Portugal y Francia son los UNICOS paises de spot_price con interconexion fisica a España.
# El resto (Alemania, Italia, Suiza, Belgica, Holanda, Austria, Polonia, Chequia) no tiene cable
# a España -- su correlacion con el precio español (0.5-0.68) probablemente ya la capturan
# gas_ttf/co2_ets, y meterlos sin un canal causal real arriesga redundancia sin justificacion.
COLS_PRECIO_VECINOS = ["pt_entsoe", "fr_entsoe"]


def _features_lag_precio_vecinos(conn) -> pd.DataFrame:
    """Precio real de Portugal y Francia, D-1 y D-7 -- NUNCA el propio D+1. FUERA del dataset por
    defecto desde 20-ago-2026 -- decision de reunion del equipo, ver
    docs/columnas_pendientes_equipo.md. Solo se activa con `incluir_columnas_pendientes=True`.
    El precio de esos paises para D+1 se fija en la MISMA subasta simultanea que el español
    (acoplamiento SDAC/MIBEL) -- no se conoce con antelacion, usarlo sin lag seria casi tan
    circular como usar el propio precio de España. Portugal correlaciona 0.997 con España
    (practicamente el mismo mercado, coincide en el 94.9% de las horas); Francia 0.70 (conexion real, con capacidad
    limitada frente al tamaño de ambos mercados)."""
    df = pd.read_sql(
        f"SELECT datetime, {', '.join(c + '::float8 AS ' + c for c in COLS_PRECIO_VECINOS)} "
        f"FROM spot_price WHERE datetime BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True)
    df["fecha"] = df["datetime"].dt.tz_convert("Europe/Madrid").dt.date

    precio_diario = df.groupby("fecha")[COLS_PRECIO_VECINOS].agg(["mean", "min", "max"])
    precio_diario.columns = [f"{c}_{stat}" for c, stat in precio_diario.columns]

    lag1 = _construir_lag(precio_diario, 1, "lag1d")
    lag7 = _construir_lag(precio_diario, 7, "lag7d")
    return lag1.join(lag7, how="outer")


# Las 9 columnas de medicion de era5_weather_agg -- ampliado 20-ago-2026 a peticion del equipo,
# antes solo se usaban 6. Quedan fuera tensor_path/tensor_index: son metadatos internos del
# pipeline de descarga (ruta del tensor NetCDF, indice), no variables fisicas.
COLS_CLIMA = ["t2m_mean", "d2m_mean", "wind10_mean", "wind100_mean", "wind_gust10_mean",
              "ssrd_mean", "tcc_mean", "tp_mean", "msl_mean"]


def _features_clima(conn) -> pd.DataFrame:
    """Clima (ERA5) agregado por dia, alineado a D+1 -- MISMO criterio de shift que
    `_features_forecast` (describe D+1, se coloca en la fila D que lo usaria).

    Incluida por defecto desde el 20-ago-2026, con visto bueno del equipo. OJO -- la salvedad de
    seguridad sigue siendo real y vale la pena tenerla presente pese a la aprobacion: ERA5 es
    reanalisis (clima REAL ya ocurrido), no una prevision. Usar el clima de D+1 asume que en
    produccion se tendria una prevision ECMWF igual de buena para D+1, lo cual todavia no esta
    validado (`ecmwf_forecast_agg` solo tiene unos dias de historico) -- pendiente de conversacion
    del equipo, ver conversacion 20-ago-2026. El equipo decidio aceptar esa asuncion por ahora
    -- para desactivarla, `incluir_clima=False`.
    """
    df = pd.read_sql(
        f"SELECT ts, {', '.join(COLS_CLIMA)} FROM era5_weather_agg WHERE ts BETWEEN %(start)s AND %(end)s ORDER BY ts",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df["ts"] = pd.to_datetime(df["ts"], utc=True)   # normalizacion UTC manual, NUNCA parse_dates
    # Kelvin -> Celsius: t2m (temperatura) y d2m (punto de rocio) son las dos unicas en Kelvin
    # (verificado 20-ago-2026: d2m_mean ronda 276-277, mismo rango que t2m_mean sin convertir).
    df["t2m_mean"] = df["t2m_mean"] - 273.15
    df["d2m_mean"] = df["d2m_mean"] - 273.15
    df["fecha_objetivo"] = df["ts"].dt.tz_convert("Europe/Madrid").dt.date  # dia D+1 que describe

    feats = df.groupby("fecha_objetivo")[COLS_CLIMA].agg(["mean", "min", "max"])
    feats.columns = [f"{c}_{stat}" for c, stat in feats.columns]
    feats.index = (pd.to_datetime(feats.index) - pd.Timedelta(days=1)).date  # -> fila del dia D
    feats.index.name = "fecha"
    return feats


def construir_dataset_diario(solo_filas_validas: bool = True, incluir_clima: bool = False,
                              incluir_columnas_pendientes: bool = False,
                              avisar_constantes: bool = True,
                              filtrar_cobertura: bool = True) -> pd.DataFrame:
    """Construye el dataset maestro completo: target D+1 + features seguras + lags reales.

    Parametros:
        solo_filas_validas: si True (por defecto), excluye filas con el target D+1 incompleto
            (cambio de hora en marzo, borde final de la ventana de datos).
        incluir_clima: DESACTIVADO POR DEFECTO desde el 23-ago-2026 (antes True). ERA5 es
            REANALISIS: el clima realmente observado el dia D+1. Como el dataset ya lleva
            gen_wind_prev_mw y gen_solar_pv_prev_mw -- que son la prevision oficial legitima de
            renovables para D+1 -- la informacion MARGINAL que aporta ERA5 es exactamente
            aquello en lo que la prevision oficial se equivoco. Es decir, el modelo aprende a
            corregir el error de prevision de REE usando el resultado, y en produccion esa
            correccion no existe. Los MAE de test con clima activo no son reproducibles.
            USO RECOMENDADO: entrenar las dos versiones y reportar la diferencia como ABLACION
            DE PREVISION METEOROLOGICA PERFECTA -- da una cota superior cuantificada del valor
            de mejorar el forecast meteo, que es el argumento del capitulo del tensor ECMWF.
            Etiquetar siempre los resultados que la usen.
        avisar_constantes: imprime al final las columnas con un solo valor distinto (p.ej. las
            `_min` de solar y termosolar, que valen 0 todas las noches). No las elimina.
        incluir_columnas_pendientes: si True, añade las 4 familias de columnas que la reunion del
            equipo del 20-ago-2026 dejo FUERA del dataset por ahora (previsión ENTSO-E, diferencia
            entre previsiones, capacidad disponible, precio de paises vecinos) -- ver
            docs/columnas_pendientes_equipo.md para el detalle y el porque. Por defecto False.

    Devuelve un DataFrame indexado por "fecha" (el dia D, dia en que se hace la prediccion).
    """
    conn = _conectar()
    try:
        _inicio_efectivo(conn)
        # target_wide SIN shift: "precio real por dia", se reutiliza para los lags de precio
        target_wide = pd.read_sql(
            "SELECT datetime, es_esios::float8 AS es_esios FROM spot_price "
            "WHERE datetime BETWEEN %(start)s AND %(end)s ORDER BY datetime",
            conn, params={"start": DATASET_START, "end": DATASET_END},
        )
        target_wide["datetime"] = pd.to_datetime(target_wide["datetime"], utc=True)
        target_wide["fecha_madrid"] = target_wide["datetime"].dt.tz_convert("Europe/Madrid").dt.date
        target_wide["hora"] = target_wide["datetime"].dt.tz_convert("Europe/Madrid").dt.hour
        target_wide = target_wide.pivot_table(index="fecha_madrid", columns="hora", values="es_esios", aggfunc="first")
        target_wide.columns = [f"price_h{h:02d}" for h in target_wide.columns]
        target_wide.index.name = "fecha"

        target = _target_d1(conn)                       # con shift -1: target real de D+1
        feats_fcst = _features_forecast(conn)
        feats_ntc = _features_ntc(conn)
        feats_dia_d = _features_dia_d(conn)
        cal = _calendario(target.index)
        feats_lag_real = _features_lag_reales(conn)
        feats_lag_precio = _features_lag_precio(target_wide)
        feats_vecinos_d0 = _features_precio_vecinos_d0(conn)
        feats_pdbc = _features_pdbc(conn)
        feats_cap_inst = _features_capacidad(conn, "esios_capacity_installed",
                                             COLS_CAP_INSTALADA, "capinst")
        feats_cap_disp = _features_capacidad(conn, "esios_capacity_available",
                                             COLS_CAP_DISPONIBLE, "capdisp")
        feats_clima = _features_clima(conn) if incluir_clima else None

        if incluir_columnas_pendientes:
            feats_fcst_entsoe = _features_forecast_entsoe(conn)
            feats_diff_previsiones = _features_diferencia_previsiones(conn)
            feats_capacidad = _features_capacidad_disponible(conn)
            feats_lag_precio_vecinos = _features_lag_precio_vecinos(conn)
        else:
            feats_fcst_entsoe = feats_diff_previsiones = feats_capacidad = feats_lag_precio_vecinos = None
    finally:
        conn.close()

    piezas = [feats_fcst, feats_ntc, feats_dia_d, cal, feats_lag_real, feats_lag_precio,
              feats_vecinos_d0]
    for pieza in (feats_pdbc, feats_cap_inst, feats_cap_disp, feats_fcst_entsoe,
                  feats_diff_previsiones, feats_capacidad, feats_lag_precio_vecinos, feats_clima):
        if pieza is not None and not (isinstance(pieza, pd.DataFrame) and pieza.empty):
            piezas.append(pieza)
    dataset = target.join(piezas, how="left")

    # Corte de modelado: la fila D predice D+1, asi que se conservan las filas cuyo TARGET
    # (D+1) no supere MODELO_END -> ultima fila util D = MODELO_END - 1 dia.
    idx_dt = pd.to_datetime(dataset.index)
    dataset = dataset[idx_dt <= pd.Timestamp(MODELO_END) - pd.Timedelta(days=1)].copy()

    if solo_filas_validas:
        target_cols = [c for c in dataset.columns if c.startswith("price_h")
                       and not c.endswith(("_d0", "_lag1d", "_lag7d"))]
        dataset = dataset[dataset[target_cols].notna().all(axis=1)].copy()

    # ELIMINADO el round(2) global (estaba aqui desde el 21-ago-2026). Motivo: aniquila
    # columnas de magnitud pequeña. tp_mean de ERA5 viene en METROS (valores ~0,001-0,005) y
    # round(2) la deja en 0,00 en todas las filas; las ciclicas seno/coseno pierden resolucion.
    # El redondeo era una peticion de PRESENTACION, asi que se aplica solo al escribir el CSV
    # (ver __main__), no al objeto que entra al modelo.

    if filtrar_cobertura:
        tgt = [c for c in dataset.columns if c.startswith("price_h")
               and not c.endswith(("_d0", "_lag1d", "_lag7d"))]
        dataset = _filtrar_por_cobertura(dataset, proteger=tgt)
    dataset = _filtrar_filas_incompletas(dataset)

    bloques = {
        "target (D+1)": [c for c in dataset.columns if c.startswith("price_h")
                         and not c.endswith(("_d0", "_lag1d", "_lag7d"))],
        "forecast":     [c for c in dataset.columns if "_prev_mw" in c or "autoconsumo" in c],
        "ntc":          [c for c in dataset.columns if c.startswith("ree_ntc_")],
        "commodities":  [c for c in dataset.columns if c.split("_")[0] in ("gas", "co2")
                         or c == "dias_desde_cierre"],
        "calendario":   [c for c in dataset.columns if c.startswith("d1_")],
        "reales lag":   [c for c in dataset.columns if c.endswith(("_lag1d", "_lag7d"))
                         and not c.startswith(("price_h", "pdbc_"))],
        "precio feat":  [c for c in dataset.columns if c.startswith("price_h")
                         and c.endswith(("_d0", "_lag1d", "_lag7d"))]
                        + [c for c in dataset.columns if c.startswith("precio_real_")],
        "vecinos":      [c for c in dataset.columns if c.endswith("_d0")
                         and not c.startswith("price_h")],
        "pdbc":         [c for c in dataset.columns if c.startswith("pdbc_")],
        "cap instalada":[c for c in dataset.columns if c.startswith("capinst_")],
        "cap disponible":[c for c in dataset.columns if c.startswith("capdisp_")],
        "clima":        [c for c in dataset.columns if c.split("_")[0] in
                         ("t2m", "d2m", "wind10", "wind100", "ssrd", "tcc", "tp", "msl")],
    }
    print("\nColumnas por bloque:")
    for nombre, cols_b in bloques.items():
        print(f"  {nombre:<16} {len(cols_b):>4}")
    print(f"  {'TOTAL':<16} {dataset.shape[1]:>4}\n")

    if avisar_constantes:
        const = [c for c in dataset.columns if dataset[c].nunique(dropna=True) <= 1]
        if const:
            print(f"AVISO: {len(const)} columnas constantes (sin poder predictivo): {const}")
        obj = list(dataset.dtypes[dataset.dtypes == "object"].index)
        if obj:
            print(f"AVISO: {len(obj)} columnas dtype=object (Decimal de postgres numeric); "
                  f"Keras/sklearn fallaran: {obj}")

    return dataset


# ══════════════════════════════════════════════════════════════════════════════════════════
#  DATASET HORARIO (formato largo) -- añadido 23-ago-2026
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# POR QUE EXISTE. El dataset diario colapsa cada exogena a mean/min/max del dia. Eso destruye
# justo la informacion que mapea uno a uno contra las 24 salidas: el valle de precio de las
# 14:00 lo explica la prevision solar DE LAS 14:00, no la media del dia. Aqui cada fila es un
# par (dia objetivo, hora) y las exogenas entran con su valor horario.
#
# TRES REGLAS QUE NO SE PUEDEN ROMPER
#
#   1) NINGUN LAG MENOR DE 24 HORAS. A las 12:00 de D se predicen las 24 horas de D+1 DE UNA
#      VEZ. Cuando se predice la hora 14 no se conoce la hora 13. Cualquier `precio_h-1`,
#      media movil intradiaria de D+1 o similar es fuga total. Todos los lags de aqui son en
#      dias completos.
#
#   2) EL SPLIT SE PARTE POR DIA, NUNCA POR FILA. Las 24 filas de un dia comparten el mismo
#      conjunto de informacion. Un split aleatorio pone horas del mismo dia a ambos lados y el
#      modelo "acierta" copiando de sus vecinas. `dividir_train_val_test` corta por fecha_pred.
#
#   3) 58.000 FILAS NO SON 58.000 OBSERVACIONES. El tamaño muestral efectivo sigue siendo el
#      numero de dias (~2.400). Se gana resolucion de features, no independencia estadistica;
#      los intervalos de confianza hay que calcularlos por dia, no por fila.
#
# CONVENCION DE NOMBRES. Todo se ancla a `fecha_objetivo` = D+1 (el dia que se predice):
#     _D    -> dia D      (dia en que se predice; su precio YA se conoce, se caso ayer)
#     _Dm1  -> dia D-1    (ultimo dia con generacion real COMPLETA a las 12:00 de D)
#     _Dm6  -> dia D-6    (mismo dia de la semana que D+1)
# Se evita a proposito el sufijo `lagNd` del dataset diario: alli el ancla es la fila D y aqui
# es D+1, asi que el mismo numero significaria cosas distintas. Mezclar las dos convenciones es
# la via mas rapida a un error de un dia.
#
# REUTILIZA el contexto diario del otro constructor (commodities, calendario, capacidad, PDBC,
# reales agregados) uniendolo por `fecha_pred` = D: son valores constantes dentro del dia, se
# repiten en las 24 filas. No se duplica logica.

# Exogenas horarias de D+1 -- las mismas de COLS_SEGURAS_FORECAST pero SIN agregar
# ══ TABLA `forecast` (24-ago-2026) ═══════════════════════════════════════════════════════
# El script leia `esios_forecast_da` (nombres tipo demanda_mercado_prev_mw, gen_wind_prev_mw).
# Pero existe la tabla `forecast` con EXACTAMENTE las 12 columnas de la matriz FINAL del equipo
# y sus nombres. Verificado contra information_schema. Se pasa a leer de ahi, y con ello:
#
#   - `c_autoconsumo_prev` VIENE DADA, no hay que reconstruirla por resta. Ojo a su signo: en
#     enero de 2020 es NEGATIVA en horas de sol (-1.285 MW) y positiva de noche. No es una
#     magnitud fisica de autoconsumo, es el residuo entre dos series -- solo interpretable como
#     autoconsumo desde diciembre de 2025.
#   - `autoconsumo_estimado` EXISTE como boolean (vale false en 2020). Es el flag del cambio
#     metodologico de REE. Se castea a int; si no tiene varianza en train (train acaba en
#     dic-2024 y el cambio es de dic-2025) el filtro lo descartara solo, y eso es correcto.
#   - Las seis NTC PREVISTAS son las que pide la matriz. Son inequivocamente EX-ANTE, a
#     diferencia de las `ree_ntc_*` de load_inter, cuyo caracter no hemos podido verificar
#     (corr 0,84 entre ambas, coincidencia exacta solo en el 48% de las horas). Usar estas
#     elimina la duda de fuga.
# Poner a True solo para comparar: mete tambien las ree_ntc_* de load_inter, cuya condicion
# ex-ante/ex-post esta sin verificar.
# ══ PRECIOS EUROPEOS (24-ago-2026) ═══════════════════════════════════════════════════════
# Las 13 zonas de spot_price ademas de es_esios. Cobertura verificada: 99,96% o mas en todas,
# asi que ninguna se cae por falta de datos -- si alguna sobra sera por redundancia, no por
# huecos.
#
# TODAS DEL DIA D, NUNCA DE D+1. Y no es una precaucion: es una imposibilidad fisica. Todas las
# zonas SDAC se casan A LA VEZ al cierre de las 12:00 de D-1 y se publican hacia las 12:45. El
# precio frances del dia D+1 no existe cuando hay que predecir el español del dia D+1 -- son el
# resultado de la misma subasta. Usarlo seria fuga por circularidad.
# El propio pipeline del equipo lo documenta: `price_fr_eur_mwh` se retiro de
# entsoe_forecast_da el 17-ago-2026 por exactamente este motivo.
#
# Van al DECODER alineadas por hora, igual que es_esios_D: al predecir la hora 14 de D+1, tener
# el precio frances de la hora 14 de D en esa misma posicion es un ancla directa.
#
# LOS SPREADS son la parte interesante. PT correlaciona 0,997 con España y coincide en el 94,9%
# de las horas, asi que como NIVEL aporta poco; pero las horas en que los dos mercados se
# DESACOPLAN son las de congestion en la interconexion, que son justo las de precio extremo y
# las que peor predice el modelo. El spread las señala; la columna cruda no.
COLS_PRECIOS_EUROPA = ["pt_entsoe", "fr_entsoe", "de_lu_entsoe", "it_nord_entsoe", "ch_entsoe",
                       "be_entsoe", "nl_entsoe", "at_entsoe", "pl_entsoe", "cz_entsoe"]
SPREADS_EUROPA = ["pt_entsoe", "fr_entsoe", "de_lu_entsoe"]   # es_esios - zona, del dia D
INCLUIR_PRECIOS_EUROPA = True

INCLUIR_NTC_LOAD_INTER = False

TABLA_FORECAST = "forecast"
COLS_FORECAST = ["ree_demanda_prev", "ree_gwind_prev", "ree_gsolar_prev",
                 "ree_grenov_prev", "ree_ntc_impfr_prev", "ree_ntc_expfr_prev",
                 "ree_ntc_imppt_prev", "ree_ntc_exppt_prev", "ree_ntc_impma_prev",
                 "ree_ntc_expma_prev"]
COL_FLAG_AUTOCONSUMO = None       # ver EXCLUIR_AUTOCONSUMO

# ══ AUTOCONSUMO: EXCLUIDO POR DECISION (24-ago-2026) ═════════════════════════════════════
# REE no incorpora la estimacion de autoconsumo a sus series hasta el 1-dic-2025. Verificado
# en la BD, y el corte es limpio: `autoconsumo_estimado` vale false desde 2020-01-01 hasta
# 2025-11-30 (51.864 horas) y true desde 2025-12-01 (6.431 horas), sin alternancias.
#
# Eso deja las dos columnas inservibles con este split (train <= 2024, validation 2025,
# test 2026), y por motivos distintos:
#
#   autoconsumo_estimado  -> CONSTANTE en train. El modelo no puede aprender nada de un flag
#       que siempre vale lo mismo mientras entrena.
#
#   c_autoconsumo_prev    -> son DOS magnitudes distintas bajo el mismo nombre. Antes de
#       dic-2025 es el residuo entre 1775 y 2563, dos series casi identicas: oscila alrededor
#       de cero con signo variable (-7,4 en sep-2025; -124,2 en nov-2025). Un "autoconsumo
#       previsto" NEGATIVO no tiene sentido fisico -- es ruido. Desde dic-2025 pasa a ser
#       autoconsumo de verdad, y no como un escalon sino como una RAMPA:
#           dic-2025  +342 MW      mar-2026  +1.582
#           ene-2026  +715         may-2026  +1.987
#           feb-2026 +1.097        jul-2026  +2.118
#       Se multiplica por seis en siete meses. Es exactamente el patron que eliminan los
#       filtros de tendencia: el modelo entrena viendo ~0 y en test se encuentra 2.145 MW,
#       fuera de todo rango aprendido. Entraria clipada a +-10 sigmas, o sea como constante.
#
# No se pierde informacion utilizable: ninguna de las dos PUEDE aportar nada dado el split.
# Lo que se gana es que la exclusion queda documentada como decision razonada, con los numeros
# delante, en vez de ocurrir de rebote dentro de un filtro automatico.
EXCLUIR_AUTOCONSUMO = True

COLS_HORARIAS_FORECAST = COLS_SEGURAS_FORECAST + COLS_SEGURAS_FORECAST_EXTRA

# Reales horarios que se traen a la misma hora de D-1 y D-6
COLS_HORARIAS_REALES_LOAD = COLS_DEMANDA_REAL
COLS_HORARIAS_REALES_ENTSOE = COLS_ENTSOE_REAL
COLS_HORARIAS_REALES_ESIOS = COLS_ESIOS_REAL


def _largo_horario(conn, tabla: str, cols: list, etiqueta: str) -> pd.DataFrame:
    """Lee una tabla horaria y la devuelve en largo: (fecha, hora, cols...) en hora de Madrid.

    El `groupby(...).mean()` no es decorativo: el dia de 25 horas de octubre tiene DOS horas
    02:00 (CEST y CET). El dataset diario las resolvia con pivot_table(aggfunc="first"), que
    descarta la segunda en silencio. Aqui se promedian, que es lo correcto para una magnitud
    horaria, y ademas colapsa cualquier duplicado de origen sin romper el merge posterior.
    """
    cols = _filtrar_existentes(conn, tabla, cols, etiqueta)
    if not cols:
        return pd.DataFrame()
    ts = _col_temporal(conn, tabla)
    cast = ", ".join(f"{c}::float8 AS {c}" for c in cols)   # numeric -> Decimal -> dtype object
    df = pd.read_sql(
        f"SELECT {ts}, {cast} FROM {tabla} WHERE {ts} BETWEEN %(start)s AND %(end)s",
        conn, params={"start": DATASET_START, "end": DATASET_END},
    )
    df[ts] = pd.to_datetime(df[ts], utc=True)             # normalizacion manual, NUNCA parse_dates
    loc = df[ts].dt.tz_convert("Europe/Madrid")
    df["fecha"] = loc.dt.date
    df["hora"] = loc.dt.hour
    return df.groupby(["fecha", "hora"], as_index=False)[cols].mean()


def _desplazar_dias(largo: pd.DataFrame, dias: int, sufijo: str) -> pd.DataFrame:
    """Adelanta la fecha `dias` dias para que el dato de X quede alineado con la fila X+dias.
    Trabaja sobre fechas naive, asi que Timedelta es seguro aqui (no hay husos que perder)."""
    out = largo.copy()
    out["fecha"] = (pd.to_datetime(out["fecha"]) + pd.Timedelta(days=dias)).dt.date
    ren = {c: f"{c}_{sufijo}" for c in out.columns if c not in ("fecha", "hora")}
    return out.rename(columns=ren)


# ══════════════════════════════════════════════════════════════════════════════════════════
#  EXPORTACION DOCUMENTADA A EXCEL (24-ago-2026)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# Dos hojas:
#   MATRIZ       datos con una fila de cabecera POR ENCIMA de los nombres, en negrita y
#                fusionada, que agrupa cada bloque por su TABLA DE ORIGEN. Las columnas
#                CALCULADAS (no vienen tal cual de ninguna tabla) se marcan en ambar.
#   EXCLUSIONES  que columnas y que fechas se han descartado y por que, con el dato de respaldo.
#
# El motivo de la segunda hoja: los filtros de tendencia, cobertura y periodo anomalo quitan
# unas 25 columnas y varios cientos de dias. Sin dejar rastro, dentro de dos meses nadie
# recuerda por que falta capinst_solar_pv_mw -- y es exactamente el tipo de cosa que pregunta
# un tribunal.

# (prefijo/sufijo -> (tabla de origen, color de fondo RGB))
ORIGEN_COLUMNAS = [
    (("fecha_pred", "fecha_objetivo", "hora"), "CLAVE", "D9D9D9"),
    (("target_price",), "spot_price (TARGET)", "FFC7CE"),
    (("es_esios_D", "spread_es_"), "spot_price (lag)", "FFE1E4"),
    (("pt_entsoe_D", "fr_entsoe_D", "de_lu_entsoe_D", "it_nord_entsoe_D", "ch_entsoe_D",
      "be_entsoe_D", "nl_entsoe_D", "at_entsoe_D", "pl_entsoe_D", "cz_entsoe_D",
      "pt_omie_D"), "spot_price (Europa, dia D)", "F8CBAD"),
    (("ree_ntc_impfr_prev", "ree_ntc_expfr_prev", "ree_ntc_imppt_prev", "ree_ntc_exppt_prev",
      "ree_ntc_impma_prev", "ree_ntc_expma_prev"), "forecast (NTC D+1)", "A9D08E"),
    (("_prev", "_prev_mw", "c_autoconsumo_prev", "autoconsumo_estimado"), "forecast", "C6E0B4"),
    (("ree_ntc_",), "load_inter (NTC)", "A9D08E"),
    # Las 21 series reales horarias van bajo UN SOLO bloque `generation`, aunque procedan de
    # tres tablas distintas (load_inter, entsoe_gen_data, esios_gen). Para leer la matriz lo
    # que importa es que son la MISMA cosa -- el estado real del sistema hora a hora, que
    # alimenta los canales del encoder -- no de que tabla se descargo cada una.
    (("entsoe_load_", "ree_load_", "ree_netflow_", "total_net_flow_", "gen_peninsular_",
      "wind_mw_", "pumping_cons_mw_", "hydro_run_river_mw_", "biomass_mw_", "waste_mw_",
      "other_renewable_mw_", "oil_mw_", "hydro_dispatch_mw_",
      "ree_g*", "ree_c*", "solar_fv_mw_"), "generation", "9DC3E6"),
    (("gas_", "co2_"), "commodities", "FFD966"),
    (("pdbc_",), "esios_pdbc_gen", "F4B183"),
    (("capinst_",), "esios_capacity_installed", "D6B4E8"),
    (("capdisp_",), "esios_capacity_available", "C9A0DC"),
    (("t2m_", "d2m_", "msl_", "wind10_", "wind100_", "wind_gust10_", "tcc_", "ssrd_",
      "tp_acum_"), "era5_weather_agg", "B7E1CD"),
    (("d1_", "hora_sin", "hora_cos"), "CALENDARIO (calculado)", "FFF2CC"),
    (("split",), "CLAVE", "D9D9D9"),
]

# Columnas que NO vienen tal cual de ninguna tabla: se calculan aqui. Se resaltan en ambar para
# que quien lea el Excel sepa que su definicion esta en este script y no en la BD.
COLUMNAS_CALCULADAS = ("c_autoconsumo_prev", "solar_fv_mw", "hydro_dispatch_mw",
                       "total_net_flow_mw", "d1_", "hora_sin", "hora_cos", "capinst_",
                       "capdisp_")


def _origen(col: str) -> tuple:
    """Tabla de origen y color de una columna.

    Reglas EXPLICITAS, no heuristicas:
        "_xxx"  -> casa por SUFIJO
        "xxx_"  -> casa por PREFIJO
        "xxx*"  -> casa por PREFIJO (para prefijos sin guion bajo, como "ree_g")
        "xxx"   -> coincidencia EXACTA
    La primera version usaba startswith para todo y clasificaba `hora_sin` como CLAVE, porque
    empieza por "hora". El orden de ORIGEN_COLUMNAS importa: "ree_ntc_" y "ree_load_" van antes
    que "ree_g*" para que no se las quede el bloque equivocado.
    """
    for claves, tabla, color in ORIGEN_COLUMNAS:
        for k in claves:
            if k.startswith("_"):                       # sufijo:  "_Dm1"
                if col.endswith(k):
                    return tabla, color
            elif k.endswith(("_", "*")):                 # prefijo: "pdbc_", "ree_g*"
                if col.startswith(k.rstrip("*")):
                    return tabla, color
            elif col == k:                               # exacto:  "hora", "split"
                return tabla, color
    return "SIN CLASIFICAR", "FFFFFF"


def _es_calculada(col: str) -> bool:
    return any(col.startswith(k) if k.endswith("_") else col == k
               for k in COLUMNAS_CALCULADAS)


def exportar_excel(df: pd.DataFrame, ruta: Path, variante: str = "", max_filas: int = None):
    """Escribe el Excel documentado. `max_filas` limita la hoja MATRIZ (None = todas).

    Con las 57.425 filas completas son ~3,7 millones de celdas. openpyxl las escribe, pero hay
    que ayudarle: los valores se convierten a tipos nativos ANTES de entrar (las fechas de
    pandas y los numpy.float64 obligan a openpyxl a inspeccionar celda por celda, y eso es lo
    que dispara el tiempo). Con la conversion previa baja de varios minutos a menos de uno.

    AVISO SOBRE EL LIMITE DE EXCEL: 1.048.576 filas. Con 2.393 dias x 24 h caben de sobra, pero
    si algun dia se pasa a resolucion de 15 minutos (4x) habria que trocear por año.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    datos = df if max_filas is None else df.head(max_filas)
    wb = Workbook()

    # ── hoja MATRIZ ───────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "MATRIZ"
    fina = Side(style="thin", color="808080")

    grupos, actual, ini = [], None, 1
    for i, c in enumerate(datos.columns, start=1):
        tabla, _ = _origen(c)
        if tabla != actual:
            if actual is not None:
                grupos.append((actual, ini, i - 1))
            actual, ini = tabla, i
    grupos.append((actual, ini, len(datos.columns)))

    for tabla, a, b in grupos:                      # fila 1: origen, negrita, fusionada
        _, color = _origen(datos.columns[a - 1])
        cel = ws.cell(row=1, column=a, value=tabla)
        cel.font = Font(name="Arial", bold=True, size=10)
        cel.fill = PatternFill("solid", fgColor=color)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = Border(left=fina, right=fina, top=fina, bottom=fina)
        if b > a:
            ws.merge_cells(start_row=1, start_column=a, end_row=1, end_column=b)
            for j in range(a + 1, b + 1):
                ws.cell(row=1, column=j).fill = PatternFill("solid", fgColor=color)
                ws.cell(row=1, column=j).border = Border(top=fina, bottom=fina, right=fina)

    ambar = PatternFill("solid", fgColor="FFC000")
    for j, c in enumerate(datos.columns, start=1):  # fila 2: nombres
        cel = ws.cell(row=2, column=j, value=c)
        cel.font = Font(name="Arial", bold=True, size=9,
                        color="7F6000" if _es_calculada(c) else "000000")
        if _es_calculada(c):
            cel.fill = ambar
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
        cel.border = Border(left=fina, right=fina, top=fina, bottom=fina)
        ws.column_dimensions[get_column_letter(j)].width = max(11, min(len(c) + 2, 22))

    # Conversion previa a tipos nativos: es lo que hace viable escribir 3,7M de celdas.
    # OJO con el orden de las comprobaciones: se pregunta primero si es NUMERICA, no si es
    # texto. En pandas 3 las columnas de texto tienen dtype `str`, no `object`, asi que un
    # `dtype == object` falla y las manda al to_numeric -> NaN -> celdas vacias. Paso: las dos
    # columnas de fecha salian en blanco en todo el Excel.
    bloque = datos.copy()
    for c in bloque.columns:
        col = bloque[c]
        if pd.api.types.is_datetime64_any_dtype(col):
            bloque[c] = col.dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_numeric_dtype(col) or pd.api.types.is_bool_dtype(col):
            bloque[c] = pd.to_numeric(col, errors="coerce")
        else:
            bloque[c] = col.astype(str).replace({"nan": None, "NaT": None, "<NA>": None})
    for fila in bloque.itertuples(index=False, name=None):
        ws.append(list(fila))

    ws.freeze_panes = "E3"
    # Autofiltro sobre la fila de nombres: con el dataset entero, poder filtrar por fecha o por
    # split es lo que hace el Excel consultable en vez de solo legible.
    ws.auto_filter.ref = f"A2:{get_column_letter(len(datos.columns))}{len(datos) + 2}"

    # ── hoja EXCLUSIONES ──────────────────────────────────────────────────────────────────
    we = wb.create_sheet("EXCLUSIONES")
    we["A1"] = f"Variables y fechas descartadas -- variante {variante}"
    we["A1"].font = Font(name="Arial", bold=True, size=12)
    we["A2"] = ("Cada fila documenta QUE se quita, POR QUE y con que dato de respaldo. Todos "
                "los criterios se miden SOLO sobre train: usar test para seleccionar variables "
                "contaminaria la evaluacion.")
    we["A2"].font = Font(name="Arial", size=9, italic=True)
    we.merge_cells("A2:D2")
    we["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    we.row_dimensions[2].height = 30

    cab = ["Tipo", "Elemento", "Motivo", "Detalle / dato de respaldo"]
    for j, t in enumerate(cab, start=1):
        cel = we.cell(row=4, column=j, value=t)
        cel.font = Font(name="Arial", bold=True, size=10)
        cel.fill = PatternFill("solid", fgColor="D9D9D9")
        cel.border = Border(left=fina, right=fina, top=fina, bottom=fina)
    col_tipo = {"columna": "FFC7CE", "fechas": "FFE699", "transformacion": "C6E0B4"}
    for i, r in enumerate(REGISTRO_EXCLUSIONES, start=5):
        for j, k in enumerate(["tipo", "elemento", "motivo", "detalle"], start=1):
            cel = we.cell(row=i, column=j, value=r.get(k, ""))
            cel.font = Font(name="Arial", size=9)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            cel.border = Border(left=fina, right=fina, top=fina, bottom=fina)
            if j == 1:
                cel.fill = PatternFill("solid", fgColor=col_tipo.get(r.get("tipo"), "FFFFFF"))
    for col, anc in zip("ABCD", (15, 34, 34, 80)):
        we.column_dimensions[col].width = anc
    we.freeze_panes = "A5"

    if len(datos) > 1_048_575:
        raise ValueError(f"{len(datos):,} filas superan el limite de Excel (1.048.576). "
                         f"Usa --filas-excel para limitar, o trocea por año.")

    wb.save(ruta)
    mb = ruta.stat().st_size / 1e6
    print(f"Excel: {ruta.name} | MATRIZ {len(datos):,}x{len(datos.columns)} | "
          f"EXCLUSIONES {len(REGISTRO_EXCLUSIONES)} entradas | "
          f"{len(grupos)} bloques | {mb:.1f} MB")
    return ruta


# ══════════════════════════════════════════════════════════════════════════════════════════
#  METEOROLOGIA ERA5 (24-ago-2026)
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# DOS COSAS VERIFICADAS EN LA BD ANTES DE ESCRIBIR NADA:
#
#   1) GRANULARIDAD TRIHORARIA. 8 registros por dia, a las 0/3/6/9/12/15/18/21. Hay que
#      interpolar 2 de cada 3 horas.
#   2) `ts` ESTA EN UTC, aunque el tipo sea `timestamp WITHOUT time zone` (naive). Comprobado
#      con el solsticio: el 21-jun-2026 el maximo de ssrd cae en ts=12:00 con 938 W/m2, y el
#      mediodia solar en la peninsula ese dia es a las 14:00 hora de Madrid -> UTC+2 encaja. Si
#      ts fuera hora local, el pico estaria dos horas desplazado y a ts=06:00 ya habria bastante
#      radiacion (hay 77 W/m2, coherente con las 08:00 locales).
#      SIN ESTA CONVERSION toda la meteorologia entraria desplazada DOS horas en verano y una en
#      invierno respecto al precio, y el modelo aprenderia la relacion sol-precio mal alineada.
#
# SE INTERPOLA EN UTC Y SE CONVIERTE AL FINAL. En hora local los puntos de anclaje se mueven con
# la estacion (en verano caen en 02/05/08/11/14/17/20/23; en invierno en 01/04/07/...) y en los
# dos dias de cambio de hora la rejilla salta. En UTC la serie es regular y continua siempre.
#
# CADA VARIABLE SE INTERPOLA COMO LE CORRESPONDE, no todas igual:
#   t2m, d2m, msl, wind10, wind100, wind_gust10, tcc -> campos suaves, interpolacion temporal.
#   ssrd  -> NO lineal. Entre las 06:00 y las 09:00 UTC la curva real es una parabola, no una
#            recta, y de noche una recta puede dar valores negativos. Se interpola en forma de
#            SENO recortado a cero, que respeta la forma del ciclo diario.
#   tp    -> es un ACUMULADO. Interpolar el valor puntual reparte lluvia ficticia. Ademas la
#            precipitacion de D+1 no mueve el precio de D+1: lo que mueve la hidraulica es el
#            acumulado de semanas. Se generan tp_acum_7d y tp_acum_30d y se descarta el valor
#            horario.
#
# MODO DE USO -- ERA5 ES REANALISIS, es decir el tiempo que REALMENTE ocurrio:
#   "lag"      dias D-1 y D-2. Se conoce a las 12:00 de D, asi que NO es fuga. Configuracion
#              titular. Aviso honesto: se espera POCA ganancia, porque ree_gwind_prev y
#              ree_gsolar_prev ya son la prevision meteorologica de REE traducida a MW, hecha
#              con modelos mejores. Donde si puede aportar es en TEMPERATURA (mueve la demanda
#              por un canal distinto al de la generacion) y en PRECIPITACION ACUMULADA (condiciona
#              la hidraulica con semanas de retraso, y eso no esta en ninguna prevision de REE).
#   "perfecto" dia D+1. ES FUGA. Solo para la ablacion "cota superior con prevision meteorologica
#              perfecta": la diferencia de MAE entre este modo y "lag" cuantifica cuanto se
#              ganaria mejorando el forecast meteo, que es el argumento del capitulo del tensor
#              ECMWF. ETIQUETAR SIEMPRE los resultados que lo usen.
#   "off"      sin meteorologia.
ERA5_MODO = "lag"                 # "lag" | "perfecto" | "off"
TABLA_ERA5 = "era5_weather_agg"

# PREVISION ANTES QUE REANALISIS (decision del equipo, 30-ago-2026)
#
# Los lags `*_met_Dm1` / `*_met_Dm2` describen el tiempo de hace uno y dos dias y salian de
# ERA5. Pero ERA5 es REANALISIS y Copernicus lo publica con unos 5 dias de retraso: a las
# 11:00 del dia D, el reanalisis de D-1 y D-2 NO EXISTE. El modelo se estaba entrenando con
# una variable que en produccion no vera nunca -- exactamente el defecto que el canal
# `*_meteo` se construyo para evitar, colado por otra puerta.
#
# Con la prevision no pasa: existe antes del momento de predecir, que es lo unico que
# importa. Asi que donde hay ECMWF manda ECMWF, en entrenamiento y en produccion por igual;
# ERA5 solo cubre lo anterior a 2024-04, que es donde no hay archivo de prevision.
#
# Las dos tablas tienen las mismas 9 columnas y las mismas unidades (medido sobre jun-ago
# 2026: t2m -0,21 K, msl +5,8 Pa, wind100 +0,73 m/s -- el sesgo de prevision conocido).
#
# Por defecto False para no cambiarle el dataset a nadie sin avisar.
ERA5_PREFERIR_ECMWF = False
TABLA_ECMWF = "ecmwf_forecast_agg"
COLS_ERA5_SUAVES = ["t2m_mean", "d2m_mean", "msl_mean", "wind10_mean", "wind100_mean",
                    "wind_gust10_mean", "tcc_mean"]
COLS_ERA5_RADIACION = ["ssrd_mean"]
COL_ERA5_LLUVIA = "tp_mean"
ERA5_ACUMULADOS = [7, 30]         # dias de acumulado de precipitacion


def _era5_horario(conn) -> pd.DataFrame:
    """ERA5 trihorario -> horario en hora de Madrid, con el tratamiento propio de cada variable."""
    if ERA5_MODO == "off":
        return pd.DataFrame()
    reales = set(_columnas_de(conn, TABLA_ERA5))
    suaves = [c for c in COLS_ERA5_SUAVES if c in reales]
    rad = [c for c in COLS_ERA5_RADIACION if c in reales]
    hay_lluvia = COL_ERA5_LLUVIA in reales
    pedidas = suaves + rad + ([COL_ERA5_LLUVIA] if hay_lluvia else [])
    if not pedidas:
        print(f"AVISO [era5]: {TABLA_ERA5} sin columnas utiles")
        return pd.DataFrame()

    ts = _col_temporal(conn, TABLA_ERA5)
    df = pd.read_sql(
        f"SELECT {ts}, {', '.join(pedidas)} FROM {TABLA_ERA5} "
        f"WHERE {ts} BETWEEN %(start)s AND %(end)s ORDER BY {ts}",
        conn, params={"start": DATASET_START, "end": DATASET_END})

    if ERA5_PREFERIR_ECMWF:
        # La prevision PISA al reanalisis donde exista, no solo lo completa por la cola:
        # el objetivo es que el modelo vea en entrenamiento lo mismo que vera al predecir.
        tiene = set(_columnas_de(conn, TABLA_ECMWF))
        cols_ec = [c for c in pedidas if c in tiene]
        fc = pd.read_sql(
            f"SELECT ts AS {ts}, {', '.join(cols_ec)} FROM {TABLA_ECMWF} "
            f"WHERE ts BETWEEN %(start)s AND %(end)s ORDER BY ts",
            conn, params={"start": DATASET_START, "end": DATASET_END})
        # `wind_gust10_mean` esta en la tabla pero viene vacia: si se dejara, borraria el
        # dato de ERA5 con nulos.
        fc = fc.dropna(axis=1, how="all")
        sustituibles = [c for c in cols_ec if c in fc.columns]
        if len(fc) and sustituibles:
            fc[ts] = pd.to_datetime(fc[ts])
            df[ts] = pd.to_datetime(df[ts])
            antes_n = df[sustituibles].notna().sum().sum()
            df = df.merge(fc[[ts] + sustituibles], on=ts, how="outer", suffixes=("", "_fc"))
            pisadas = 0
            for c in sustituibles:
                nuevo = df[f"{c}_fc"]
                pisadas += int((nuevo.notna() & df[c].notna()).sum())
                df[c] = nuevo.where(nuevo.notna(), df[c])
            df = df.drop(columns=[f"{c}_fc" for c in sustituibles]).sort_values(ts)
            print(f"[era5] PREVISION ANTES QUE REANALISIS: {len(fc):,} registros de "
                  f"{TABLA_ECMWF} · {pisadas:,} celdas de ERA5 sustituidas de "
                  f"{antes_n:,} · columnas {sorted(sustituibles)}")
            fuera = sorted(set(pedidas) - set(sustituibles))
            if fuera:
                print(f"[era5]   siguen saliendo de ERA5 (ECMWF no las tiene): {fuera}")

    # naive -> UTC explicito. Es el pendiente del 17-ago sobre el que avisan las notas del
    # equipo: mezclar naive y aware en un join falla en silencio.
    df[ts] = pd.to_datetime(df[ts])
    if getattr(df[ts].dt, "tz", None) is None:
        df[ts] = df[ts].dt.tz_localize("UTC")
    df = df.set_index(ts).sort_index()

    rejilla = pd.date_range(df.index.min(), df.index.max(), freq="h", tz="UTC")
    out = pd.DataFrame(index=rejilla)

    for c in suaves:
        out[c] = df[c].reindex(rejilla.union(df.index)).interpolate("time").reindex(rejilla)

    for c in rad:
        # La radiacion sigue un ciclo diario con forma de seno. Se interpola en el tiempo y se
        # recorta a cero: una interpolacion lineal pura genera negativos al amanecer/anochecer,
        # que no existen fisicamente.
        v = df[c].reindex(rejilla.union(df.index)).interpolate("time").reindex(rejilla)
        out[c] = v.clip(lower=0)

    if hay_lluvia:
        # tp es acumulado: no se interpola el valor puntual. Se reparte el acumulado del tramo
        # y se generan ventanas moviles, que es la escala a la que la lluvia mueve la hidraulica.
        tp = df[COL_ERA5_LLUVIA].reindex(rejilla.union(df.index)).interpolate("time").reindex(rejilla)
        tp = tp.clip(lower=0)
        for d in ERA5_ACUMULADOS:
            out[f"tp_acum_{d}d"] = tp.rolling(24 * d, min_periods=1).sum()

    loc = out.index.tz_convert("Europe/Madrid")
    out = out.reset_index(drop=True)
    out["fecha"], out["hora"] = loc.date, loc.hour
    out = out.groupby(["fecha", "hora"], as_index=False).mean()
    print(f"[era5] {len(out):,} horas x {out.shape[1]-2} variables "
          f"(trihorario interpolado, UTC -> Madrid) | modo: {ERA5_MODO}")
    return out


def construir_dataset_horario(incluir_clima: bool = False,
                              incluir_contexto_diario: bool = True,
                              filtrar_cobertura: bool = True,
                              variante: str = "sin_ntc_prev") -> pd.DataFrame:
    """Dataset horario en formato largo: una fila por (dia objetivo D+1, hora).

    Devuelve columnas `fecha_objetivo` (D+1, el dia predicho), `hora` (0-23), `fecha_pred`
    (D, el dia en que se emite la prediccion y clave del split), `target_price`, y las features.

    IMPORTANTE sobre las series historicas. Con PODAR_REDUNDANTES=True solo se escriben las
    `*_Dm1` (dia D-1 hora a hora). La serie continua completa se reconstruye desde ahi: el valor
    `*_Dm1` de la fila con fecha_objetivo = T es el dato real del dia T-2, asi que reindexando
    por esa fecha se recupera la serie horaria entera y con ella cualquier ventana. Las `*_Dm6`
    que habia antes eran esa misma serie escrita otra vez con otro desplazamiento.

    En el notebook:
        hist = df[["fecha_objetivo", "hora"] + cols_dm1].copy()
        hist["fecha"] = hist["fecha_objetivo"] - pd.Timedelta(days=2)
        # pivot por (fecha, hora) -> (n_dias, 24, C) -> ventana de 168 h para el encoder

    Para pasarlo a tensor (n_dias, 24, k) del Seq2Seq basta con:
        piv = df.pivot(index="fecha_pred", columns="hora", values=<lista de features>)
    o, mas comodo, `tensor_desde_horario(df)` mas abajo.
    """
    conn = _conectar()
    try:
        _inicio_efectivo(conn)
        # ── base: el target, precio de cada (D+1, hora) ────────────────────────────────────
        precio = _largo_horario(conn, "spot_price", ["es_esios"], "precio")
        base = precio.rename(columns={"fecha": "fecha_objetivo", "es_esios": "target_price"})

        # ── exogenas horarias que DESCRIBEN D+1 (publicadas antes del cierre) ──────────────
        if DEMANDA_REAL != "ambas":
            _anotar("columna", "ree_load" if DEMANDA_REAL == "entsoe_load" else "entsoe_load",
                    "Serie de demanda duplicada",
                    "Hasta nov-2025 ree_load y entsoe_load difieren en +-4 MW (son la misma "
                    "serie); desde dic-2025 divergen hasta 2.495 MW porque ree_load incorpora "
                    "la estimacion de autoconsumo. Se conserva entsoe_load por ser "
                    "metodologicamente estable en todo el historico y estar validada con "
                    "cinco pruebas de balance del sistema")

        # Se lee de `forecast`, que trae las columnas de la matriz con sus nombres.
        cols_f = list(COLS_FORECAST)
        if not EXCLUIR_AUTOCONSUMO:
            cols_f = cols_f[:1] + ["c_autoconsumo_prev"] + cols_f[1:]
        else:
            _anotar("columna", "c_autoconsumo_prev",
                    "Cambio metodologico fuera de train",
                    "REE incorpora la estimacion de autoconsumo el 1-dic-2025. Antes es el "
                    "residuo 1775-2563 y oscila alrededor de cero con signo variable "
                    "(-124 MW en nov-2025); despues crece de 342 a 2.145 MW en 7 meses. "
                    "Train acaba en dic-2024, asi que el modelo solo veria el regimen antiguo "
                    "y en test se encontraria valores fuera de todo rango aprendido")
            _anotar("columna", "autoconsumo_estimado",
                    "Constante en train",
                    "false de 2020-01-01 a 2025-11-30 (51.864 h), true desde 2025-12-01 "
                    "(6.431 h). Train acaba en dic-2024: el flag nunca cambia mientras el "
                    "modelo entrena")
        fcst = _largo_horario(conn, TABLA_FORECAST, cols_f, "forecast")

        # el flag es boolean: se castea a int aparte, porque _largo_horario promedia
        reales_f = set(_columnas_de(conn, TABLA_FORECAST))
        if COL_FLAG_AUTOCONSUMO and COL_FLAG_AUTOCONSUMO in reales_f:
            fl = pd.read_sql(
                f"SELECT datetime, {COL_FLAG_AUTOCONSUMO}::int AS {COL_FLAG_AUTOCONSUMO} "
                f"FROM {TABLA_FORECAST} WHERE datetime BETWEEN %(start)s AND %(end)s",
                conn, params={"start": DATASET_START, "end": DATASET_END})
            fl["datetime"] = pd.to_datetime(fl["datetime"], utc=True)
            loc = fl["datetime"].dt.tz_convert("Europe/Madrid")
            fl["fecha"], fl["hora"] = loc.dt.date, loc.dt.hour
            fcst = fcst.merge(
                fl.groupby(["fecha", "hora"], as_index=False)[[COL_FLAG_AUTOCONSUMO]].max(),
                on=["fecha", "hora"], how="left")
        elif COL_FLAG_AUTOCONSUMO:
            print(f"AVISO: no existe `{COL_FLAG_AUTOCONSUMO}` en {TABLA_FORECAST}")
        # Las NTC del decoder salen de `forecast` (previstas, ex-ante). Las `ree_ntc_*` de
        # load_inter quedan FUERA: no son la misma serie (corr 0,84, coincidencia exacta en el
        # 48% de las horas) y no hemos podido verificar si son ex-post -- si lo fueran, estarian
        # describiendo D+1 con informacion posterior al cierre, es decir fuga.
        ntc = pd.DataFrame()
        if INCLUIR_NTC_LOAD_INTER:
            ntc = _largo_horario(conn, "load_inter", COLS_NTC, "ntc load_inter")

        # ── precios europeos del dia D, alineados por hora ────────────────────────────────
        pr_eu = None
        if INCLUIR_PRECIOS_EUROPA:
            cols_eu = _filtrar_existentes(conn, "spot_price", COLS_PRECIOS_EUROPA, "precios EU")
            if cols_eu:
                eu = _largo_horario(conn, "spot_price", ["es_esios"] + cols_eu, "precios EU")
                for z in SPREADS_EUROPA:
                    if z in eu.columns:
                        eu[f"spread_es_{z.split('_')[0]}"] = eu["es_esios"] - eu[z]
                eu = eu.drop(columns=["es_esios"])          # ya viaja como es_esios_D
                pr_eu = _desplazar_dias(eu, 1, "D")          # dia D -> fila del dia objetivo

        # ── precio del dia D a la misma hora ──────────────────────────────────────────────
        # Solo D. Dm1 y Dm6 serian `target_price` desplazado, es decir el mismo dato escrito
        # otra vez; el notebook los reconstruye con un shift cuando arma la ventana de 168 h.
        # D si se conserva porque va al DECODER alineado por hora, no al encoder.
        pr_D = _desplazar_dias(precio, 1, "D")
        pr_Dm1 = None if PODAR_REDUNDANTES else _desplazar_dias(precio, 2, "Dm1")
        pr_Dm6 = None if PODAR_REDUNDANTES else _desplazar_dias(precio, 7, "Dm6")

        # ── reales a la misma hora de D-1 (ultimo dia completo) y D-6 ─────────────────────
        reales = []
        for tabla, cols, et in [("load_inter", COLS_HORARIAS_REALES_LOAD, "real load"),
                                ("entsoe_gen_data", COLS_HORARIAS_REALES_ENTSOE, "real entsoe"),
                                ("esios_gen", COLS_HORARIAS_REALES_ESIOS, "real esios")]:
            largo = _largo_horario(conn, tabla, cols, et)
            if not largo.empty:
                reales.append(largo)
        # Generacion DERIVADA que faltaba en la ruta horaria (detectado 24-ago-2026):
        #   solar_fv_mw  = entsoe.solar_mw - esios.ree_gsolter_mw. ENTSO-E agrupa FV y
        #     termosolar en B16; ESIOS publica la termosolar por separado, asi que la FV limpia
        #     sale por resta. Sin ella el encoder tenia termosolar pero NO fotovoltaica, que es
        #     el principal driver del precio de mediodia. GREATEST(0,...) corrige el ~1% de
        #     horas con resta negativa por ruido de redondeo entre fuentes.
        #   hydro_dispatch_mw = embalse + turbinacion de bombeo. COALESCE solo en el bombeo: si
        #     el embalse es NULL el resultado se queda NULL, no se inventa un cero.
        for sql, cols, et in [
            ("SELECT e.datetime, GREATEST(0, e.solar_mw - s.ree_gsolter_mw) AS solar_fv_mw "
             "FROM entsoe_gen_data e JOIN esios_gen s ON e.datetime = s.datetime "
             "WHERE e.datetime BETWEEN %(start)s AND %(end)s "
             "AND e.solar_mw IS NOT NULL AND s.ree_gsolter_mw IS NOT NULL",
             ["solar_fv_mw"], "solar FV"),
            ("SELECT datetime, hydro_reservoir_mw + COALESCE(pumping_gen_mw, 0) "
             "AS hydro_dispatch_mw FROM entsoe_gen_data "
             "WHERE datetime BETWEEN %(start)s AND %(end)s",
             ["hydro_dispatch_mw"], "hidraulica despachable")]:
            try:
                d_ = pd.read_sql(sql, conn, params={"start": DATASET_START, "end": DATASET_END})
                d_["datetime"] = pd.to_datetime(d_["datetime"], utc=True)
                loc = d_["datetime"].dt.tz_convert("Europe/Madrid")
                d_["fecha"], d_["hora"] = loc.dt.date, loc.dt.hour
                reales.append(d_.groupby(["fecha", "hora"], as_index=False)[cols].mean())
            except Exception as e:
                print(f"AVISO [{et}]: no se pudo calcular ({e})")

        real = reales[0]
        for otro in reales[1:]:
            real = real.merge(otro, on=["fecha", "hora"], how="outer")

        # ── meteorologia ERA5 (ver nota en ERA5_MODO) ─────────────────────────────────────
        era5 = _era5_horario(conn)
        met_lag = met_fut = None
        if not era5.empty:
            if ERA5_MODO == "lag":
                # D-1 y D-2 (relativo a D+1: desplazar 2 y 3 dias). Se conocen a las 12:00 de D.
                met_lag = _desplazar_dias(era5, 2, "met_Dm1")
                met_lag = met_lag.merge(_desplazar_dias(era5, 3, "met_Dm2"),
                                        on=["fecha", "hora"], how="outer")
            elif ERA5_MODO == "perfecto":
                # Dia D+1: el tiempo REAL del dia que se predice. ES FUGA -- solo para la
                # ablacion de cota superior. Los nombres llevan _METEO_PERFECTA para que
                # cualquiera que abra el CSV vea de inmediato que ese resultado esta etiquetado.
                met_fut = era5.rename(columns={c: f"{c}_METEO_PERFECTA"
                                               for c in era5.columns
                                               if c not in ("fecha", "hora")})
                _anotar("transformacion", "ERA5 modo 'perfecto'",
                        "FUGA DELIBERADA -- ablacion etiquetada",
                        "ERA5 es reanalisis: el tiempo que REALMENTE ocurrio el dia D+1. No es "
                        "utilizable en produccion. La diferencia de MAE frente al modo 'lag' "
                        "cuantifica cuanto se ganaria con prevision meteorologica perfecta")
        # El dia D NO entra: a las 12:00 de D solo han ocurrido sus horas 00-11, meterlo entero
        # seria fuga de medio dia. D-1 es la ultima jornada real cerrada.
        # Dm6 tampoco se escribe: es la MISMA serie que Dm1 con otro desplazamiento, y el
        # notebook la obtiene reindexando por fecha al construir la ventana del encoder.
        re_Dm1 = _desplazar_dias(real, 2, "Dm1")
        re_Dm6 = None if PODAR_REDUNDANTES else _desplazar_dias(real, 7, "Dm6")

        # ── contexto diario: constante dentro del dia, se repite en las 24 filas ───────────
        ctx = None
        if incluir_contexto_diario:
            trozos = [_features_dia_d(conn),
                      _features_pdbc(conn),
                      _features_capacidad(conn, "esios_capacity_installed",
                                          COLS_CAP_INSTALADA, "capinst"),
                      _features_capacidad(conn, "esios_capacity_available",
                                          COLS_CAP_DISPONIBLE, "capdisp")]
            if incluir_clima:
                trozos.append(_features_clima(conn))
            trozos = [t for t in trozos if t is not None and not t.empty]
            ctx = _filtrar_tendencias(
                _normalizar_capacidad(trozos[0].join(trozos[1:], how="outer")))
    finally:
        conn.close()

    df = base
    for pieza in (fcst, ntc, met_fut):
        if pieza is None or pieza.empty:
            continue
        if True:
            df = df.merge(pieza.rename(columns={"fecha": "fecha_objetivo"}),
                          on=["fecha_objetivo", "hora"], how="left")
    for pieza in (pr_D, pr_eu, pr_Dm1, pr_Dm6, re_Dm1, re_Dm6, met_lag):
        if pieza is None:
            continue
        df = df.merge(pieza.rename(columns={"fecha": "fecha_objetivo"}),
                      on=["fecha_objetivo", "hora"], how="left")

    df["fecha_pred"] = (pd.to_datetime(df["fecha_objetivo"]) - pd.Timedelta(days=1)).dt.date

    if ctx is not None:
        ctx = ctx.reset_index().rename(columns={"fecha": "fecha_pred"})
        df = df.merge(ctx, on="fecha_pred", how="left")

    # ── calendario del dia objetivo + hora ────────────────────────────────────────────────
    cal = _calendario(sorted(df["fecha_pred"].unique()))
    cal = cal.reset_index().rename(columns={"index": "fecha_pred"})
    cal.columns = ["fecha_pred"] + list(cal.columns[1:])
    df = df.merge(cal, on="fecha_pred", how="left")

    # la hora, ciclica: para un arbol da igual, para una red 23 y 0 tienen que estar juntas
    df["hora_sin"] = np.sin(2 * np.pi * df["hora"] / 24)
    df["hora_cos"] = np.cos(2 * np.pi * df["hora"] / 24)

    # ── corte de modelado y limpieza ──────────────────────────────────────────────────────
    df = df[pd.to_datetime(df["fecha_objetivo"]) <= pd.Timestamp(MODELO_END)]
    # Exigir target es lo correcto para ENTRENAR: una fila sin precio no ensena nada. Pero
    # es justo lo contrario de lo que necesita PREDECIR: el precio de manana no existe
    # todavia -- es lo que se quiere predecir -- y con este filtro esa fila se descarta
    # siempre, se ponga lo que se ponga en MODELO_END.
    # Por defecto queda como estaba; solo `construir_matriz_produccion.py` lo pone a False.
    if EXIGIR_TARGET:
        df = df[df["target_price"].notna()]
    df = df.sort_values(["fecha_objetivo", "hora"]).reset_index(drop=True)

    fuera = [c for c in COLS_EXCLUIDAS if c in df.columns]
    if fuera:
        df = df.drop(columns=fuera)
        print(f"excluidas por COLS_EXCLUIDAS: {fuera}")

    # ── periodos anomalos: se descartan las filas y se deja constancia en el log ──────────
    fo = pd.to_datetime(df["fecha_objetivo"])
    for ini, fin in PERIODOS_EXCLUIDOS:
        m = (fo >= pd.Timestamp(ini)) & (fo <= pd.Timestamp(fin))
        if m.any():
            print(f"periodo excluido {ini} -> {fin}: {int(m.sum())} filas "
                  f"({fo[m].dt.date.nunique()} dias) descartadas")
            _anotar("fechas", f"{ini} -> {fin}", "Apagon iberico y reposicion",
                    f"{int(m.sum())} filas / {fo[m].dt.date.nunique()} dias. No es una "
                    f"observacion extrema que el modelo deba aprender: es una observacion de "
                    f"OTRO proceso (cero de suministro, restricciones masivas, arranque "
                    f"escalonado). El precio no se formo por el mecanismo habitual")
            df, fo = df[~m], fo[~m]

    orden = ["fecha_pred", "fecha_objetivo", "hora", "target_price"]
    df = df[orden + [c for c in df.columns if c not in orden]]

    # ── VARIANTE del bloque ntc_*_prev_mw (ver nota en las constantes) ────────────────────
    presentes = [c for c in COLS_NTC_PREV if c in df.columns]
    if variante == "sin_ntc_prev" and presentes:
        df = df.drop(columns=presentes)
        print(f"\n[variante sin_ntc_prev] {len(presentes)} columnas eliminadas, "
              f"se conservan todos los dias")
    elif variante == "sin_2020" and presentes:
        antes = len(df)
        hay = df[presentes].notna().any(axis=1)
        f_out = pd.to_datetime(df.loc[~hay, "fecha_objetivo"])
        df = df[hay]
        print(f"\n[variante sin_2020] columnas conservadas; {antes - len(df)} filas "
              f"({f_out.dt.date.nunique()} dias) descartadas por no tener ntc_*_prev_mw")
        if len(f_out):
            print(f"   rango descartado: {f_out.min().date()} -> {f_out.max().date()}")
    elif variante not in VARIANTES:
        raise ValueError(f"variante debe ser una de {VARIANTES}, recibido {variante!r}")

    if filtrar_cobertura:
        df = _filtrar_por_cobertura(df, proteger=orden)
    df = _filtrar_filas_incompletas(df, col_fecha="fecha_objetivo")

    bloq = {
        "clave/target": ["fecha_pred", "fecha_objetivo", "hora", "target_price"],
        "decoder D+1": [c for c in df.columns
                        if c.endswith(("_prev_mw", "_prev")) or c.startswith("ree_ntc_")
                        or c == "es_esios_D" or c.startswith("spread_es_")
                        or c.endswith(("_entsoe_D", "_omie_D"))
                        or c in ("hora_sin", "hora_cos")],
        "encoder (generation _Dm1)": [c for c in df.columns if c.endswith("_Dm1")],
        "pdbc": [c for c in df.columns if c.startswith("pdbc_")],
        "precios EU": [c for c in df.columns if c.endswith(("_entsoe_D", "_omie_D"))
                       or c.startswith("spread_es_")],
        "capacidad": [c for c in df.columns if c.startswith(("capinst_", "capdisp_"))],
        "commodities": [c for c in df.columns if c.split("_")[0] in ("gas", "co2")
                        or c == "dias_desde_cierre"],
        "calendario": [c for c in df.columns if c.startswith("d1_")],
    }
    print("\nColumnas por bloque:")
    for k, v in bloq.items():
        print(f"  {k:<24} {len(v):>4}")
    otras = set(df.columns) - {c for v in bloq.values() for c in v}
    if otras:
        print(f"  {'sin clasificar':<24} {len(otras):>4}  {sorted(otras)}")
    print(f"\ndataset horario: {df.shape[0]} filas x {df.shape[1]} columnas "
          f"({df['fecha_pred'].nunique()} dias)")
    const = [c for c in df.columns if df[c].nunique(dropna=True) <= 1]
    if const:
        print(f"AVISO: {len(const)} columnas constantes: {const}")
    obj = list(df.dtypes[df.dtypes == "object"].index)
    obj = [c for c in obj if c not in ("fecha_pred", "fecha_objetivo")]
    if obj:
        print(f"AVISO: {len(obj)} columnas dtype=object: {obj}")
    return df


def tensor_desde_horario(df: pd.DataFrame, cols_features: list):
    """Reshape del dataset horario a los tensores del Seq2Seq.

    Devuelve (X, y, fechas):
        X       (n_dias, 24, len(cols_features))
        y       (n_dias, 24)
        fechas  array de fecha_pred, en el mismo orden que la primera dimension

    Solo se conservan los dias con las 24 horas completas: un dia incompleto (cambio de hora de
    marzo, hueco de ingesta) rompe la forma del tensor. Se avisa de cuantos se descartan.
    """
    completos = df.groupby("fecha_pred")["hora"].nunique()
    validos = set(completos[completos == 24].index)
    descartados = len(completos) - len(validos)
    if descartados:
        print(f"tensor: {descartados} dias descartados por no tener 24 horas completas")
    d = df[df["fecha_pred"].isin(validos)].sort_values(["fecha_pred", "hora"])

    fechas = np.array(sorted(validos))
    n = len(fechas)
    X = d[cols_features].to_numpy(dtype="float32").reshape(n, 24, len(cols_features))
    y = d["target_price"].to_numpy(dtype="float32").reshape(n, 24)
    return X, y, fechas


# ══════════════════════════════════════════════════════════════════════════════════════════
#  TENSORES ENCODER / DECODER  -- añadido 23-ago-2026
# ══════════════════════════════════════════════════════════════════════════════════════════
#
# Separacion past covariates / future covariates. Tres bloques, y la regla de reparto NO es el
# tipo de dato sino la FRONTERA DE FUGA (12:00 del dia D, cuando cierra el mercado de D+1):
#
#   X_hist  (n, 24*V, m)  lo REALIZADO y ya conocido -> dias D-V .. D-1, hora a hora.
#                         El dia D NO entra: a las 12:00 solo han ocurrido sus horas 00-11,
#                         meterlo entero seria fuga de medio dia.
#   X_fut   (n, 24, f)    lo PUBLICADO POR ADELANTADO sobre D+1 -> previsiones y NTC horarias,
#                         MAS el precio del dia D alineado por hora, MAS hora ciclica.
#   X_est   (n, j)        lo CONSTANTE dentro del dia -> commodities D-1, capacidad, calendario,
#                         dummies de regimen.
#   y       (n, 24)       precio de D+1.
#
# POR QUE EL PRECIO DE D VA AL DECODER Y NO AL ENCODER. Es legitimo (se caso a las 12:00 de
# D-1), y podria ir al final de la secuencia del encoder. Pero puesto en X_fut queda ALINEADO
# POR HORA con la salida: al predecir la hora 14 de D+1 el decoder tiene en esa misma posicion
# el precio de la hora 14 de D. Es un ancla directa por paso de tiempo, mucho mas util que
# enterrado 24 posiciones atras en una secuencia de 168.
#
# POR QUE ESTO ARREGLA LOS NULOS. En el dataset horario las `_Dm1`/`_Dm6` son ~150 columnas
# planas, cada una con su patron de huecos al principio de la serie. Aqui son POSICIONES de una
# ventana continua: el problema se reduce a un unico calentamiento de V dias que se descarta de
# golpe. Con V=7 y datos desde 2020-01-01, el primer dia predecible es 2020-01-09.
#
# CUIDADO AL USARLO: NO son dos modelos separados cuyas predicciones se promedian. Son dos
# entradas del MISMO modelo que se concatenan antes de la salida -- la interaccion es donde
# esta la señal (el efecto de 5.000 MW de eolica prevista depende del nivel de gas y del
# regimen de precios de la semana). En Keras, dos Input() y un Concatenate.

COLS_TENSOR_REAL = (["es_esios"] + COLS_DEMANDA_REAL + COLS_ENTSOE_REAL + COLS_ESIOS_REAL)
COLS_TENSOR_FUT = COLS_SEGURAS_FORECAST + COLS_NTC


def _panel_dia_hora(largo: pd.DataFrame, cols: list, dias: pd.DatetimeIndex) -> np.ndarray:
    """(fecha, hora, cols...) -> array (n_dias, 24, len(cols)), rejilla completa.

    Reindexa contra el calendario COMPLETO y las 24 horas: el dia de 23 horas de marzo deja un
    hueco en la hora 2 que se interpola (7 horas en 6 años); el de 25 horas ya venia colapsado
    por el groupby(fecha, hora).mean() de `_largo_horario`, que promedia las dos 02:00 en vez de
    descartar una en silencio.
    """
    idx = pd.MultiIndex.from_product([dias.date, range(24)], names=["fecha", "hora"])
    piv = largo.set_index(["fecha", "hora"])[cols].reindex(idx)
    piv = piv.interpolate(limit=2, limit_direction="both")
    return piv.to_numpy(dtype="float32").reshape(len(dias), 24, len(cols))


def construir_tensores(ventana_dias: int = 7, incluir_clima: bool = False,
                       rellenar_ceros: bool = True):
    """Construye (X_hist, X_fut, X_est, y, fechas) para el modelo encoder-decoder.

    Parametros:
        ventana_dias: longitud del encoder. 7 (=168 h) captura el ciclo semanal completo y es
            el valor por defecto. 14 da mas contexto de tendencia pero dobla el coste y, con
            ~1.800 dias de train, empieza a ser mucho parametro para poco dato.
        rellenar_ceros: sustituye por 0 los NaN que quedan tras el calentamiento. Son NULOS
            ESTRUCTURALES, no datos perdidos: los indicadores de bateria (ESIOS 2166/2167) y las
            columnas de capacidad hibrida no publicaban en 2020 porque esa tecnologia no estaba
            conectada. Ahi un 0 es el valor fisicamente correcto, no una imputacion. Se avisa de
            cuantos se rellenan por bloque. IMPRESCINDIBLE para Keras: un solo NaN en X propaga
            y la perdida sale `nan` desde la primera epoca, sin lanzar ningun error.

    `fechas` es el vector de `fecha_pred` (dia D) en el mismo orden que la primera dimension;
    usalo con `dividir_tensores` para partir por dia.
    """
    conn = _conectar()
    try:
        _inicio_efectivo(conn)
        largo_real = None
        for tabla, cols, et in [("spot_price", ["es_esios"], "tensor precio"),
                                ("load_inter", COLS_DEMANDA_REAL, "tensor load"),
                                ("entsoe_gen_data", COLS_ENTSOE_REAL, "tensor entsoe"),
                                ("esios_gen", COLS_ESIOS_REAL, "tensor esios")]:
            trozo = _largo_horario(conn, tabla, cols, et)
            if trozo.empty:
                continue
            largo_real = trozo if largo_real is None else largo_real.merge(
                trozo, on=["fecha", "hora"], how="outer")

        largo_fcst = _largo_horario(conn, "esios_forecast_da", COLS_SEGURAS_FORECAST, "tensor fcst")
        largo_ntc = _largo_horario(conn, "load_inter", COLS_NTC, "tensor ntc")
        largo_fut = largo_fcst.merge(largo_ntc, on=["fecha", "hora"], how="outer")

        est = [_features_dia_d(conn),
               _features_pdbc(conn),
               _features_capacidad(conn, "esios_capacity_installed", COLS_CAP_INSTALADA, "capinst"),
               _features_capacidad(conn, "esios_capacity_available", COLS_CAP_DISPONIBLE, "capdisp")]
        if incluir_clima:
            est.append(_features_clima(conn))
        est = [e for e in est if e is not None and not e.empty]
    finally:
        conn.close()

    dias = pd.date_range(DATASET_START, MODELO_END, freq="D")

    # filtro de canales por cobertura DENTRO de la ventana efectiva (no sobre toda la serie)
    def _canales_validos(largo, cols, etiqueta):
        v = largo[(pd.to_datetime(largo["fecha"]) >= pd.Timestamp(DATASET_START))
                  & (pd.to_datetime(largo["fecha"]) <= pd.Timestamp(MODELO_END))]
        cob = v[cols].notna().mean()
        fuera, cero = [], []
        for c in cols:
            if cob[c] == 0:
                fuera.append(c)
            elif cob[c] < COBERTURA_MINIMA:
                (cero if _es_cero_estructural(c) else fuera).append(c)
        if cero:
            print(f"[{etiqueta}] {len(cero)} canales a 0 (nulo estructural): {cero}")
        if fuera:
            print(f"[{etiqueta}] {len(fuera)} canales descartados por cobertura "
                  f"< {COBERTURA_MINIMA:.0%}: "
                  + ", ".join(f"{c} ({cob[c]:.0%})" for c in fuera))
        return [c for c in cols if c not in fuera]

    cols_real = _canales_validos(largo_real,
                                 [c for c in COLS_TENSOR_REAL if c in largo_real.columns],
                                 "X_hist")
    cols_fut = _canales_validos(largo_fut,
                                [c for c in COLS_TENSOR_FUT if c in largo_fut.columns],
                                "X_fut")
    if "es_esios" not in cols_real:   # el precio nunca se descarta: es el ancla del encoder
        cols_real = ["es_esios"] + cols_real

    RE = _panel_dia_hora(largo_real, cols_real, dias)      # (T, 24, m)
    FC = _panel_dia_hora(largo_fut, cols_fut, dias)        # (T, 24, f)
    PR = RE[:, :, cols_real.index("es_esios")]             # (T, 24)

    # estaticos indexados por dia D
    est_df = _filtrar_tendencias(_normalizar_capacidad(est[0].join(est[1:], how="outer")))
    cal = _calendario(pd.to_datetime(dias).date)
    est_df = cal.join(est_df, how="left")
    est_df = est_df.reindex(pd.Index(dias.date, name="fecha"))
    est_df = _filtrar_por_cobertura(est_df, proteger=[c for c in est_df.columns
                                                      if c.startswith("d1_")])
    cols_est = list(est_df.columns)
    ES = est_df.to_numpy(dtype="float32")                  # (T, j)

    V = ventana_dias
    t0 = V + 1                    # primer indice de dia OBJETIVO con ventana completa
    t_idx = np.arange(t0, len(dias))

    X_hist = np.stack([RE[t - V - 1:t - 1].reshape(24 * V, len(cols_real)) for t in t_idx])
    y = PR[t_idx]

    hora = np.arange(24, dtype="float32")
    ciclo = np.stack([np.sin(2 * np.pi * hora / 24), np.cos(2 * np.pi * hora / 24)], axis=-1)
    X_fut = np.concatenate([
        FC[t_idx],                                  # previsiones de D+1
        PR[t_idx - 1][:, :, None],                  # precio del dia D, alineado por hora
        np.broadcast_to(ciclo, (len(t_idx), 24, 2)),
    ], axis=-1).astype("float32")
    cols_fut_full = cols_fut + ["precio_D_misma_hora", "hora_sin", "hora_cos"]

    X_est = ES[t_idx - 1]                           # estaticos del dia D
    fechas = np.array([dias.date[t - 1] for t in t_idx])   # fecha_pred = D

    # dias dentro de un periodo anomalo (apagon), y ademas las ventanas que lo pisen
    excl = np.zeros(len(dias), dtype=bool)
    for ini, fin in PERIODOS_EXCLUIDOS:
        excl |= (dias >= pd.Timestamp(ini)) & (dias <= pd.Timestamp(fin))
    ok_ventana = np.array([not excl[t - V - 1:t].any() for t in t_idx])
    if (~ok_ventana).any():
        print(f"Descartados {int((~ok_ventana).sum())} dias por periodo anomalo o por tener la "
              f"ventana del encoder solapada con el")

    # dias sin el bloque exigido (PDBC): se descarta el dia, no se mueve el inicio
    ok_bloque = np.ones(len(t_idx), dtype=bool)
    for pref in EXIGIR_BLOQUES:
        ic = [i for i, c in enumerate(cols_est) if c.startswith(pref)]
        if not ic:
            continue
        tiene = ~np.isnan(ES[t_idx - 1][:, ic]).all(axis=1)
        n = int((~tiene).sum())
        if n:
            f = pd.to_datetime([dias.date[t - 1] for t in t_idx[~tiene]])
            print(f"Descartados {n} dias sin `{pref}*` (de {f.min().date()} a {f.max().date()})")
        ok_bloque &= tiene

    # filas sin target utilizable
    ok = (~np.isnan(y).any(axis=1)) & ok_bloque & ok_ventana
    X_hist, X_fut, X_est, y, fechas = X_hist[ok], X_fut[ok], X_est[ok], y[ok], fechas[ok]

    print(f"\nTensores (ventana {V} dias = {24*V} h)")
    print(f"  X_hist {X_hist.shape}  {len(cols_real)} canales reales")
    print(f"  X_fut  {X_fut.shape}  {len(cols_fut_full)} canales futuros")
    print(f"  X_est  {X_est.shape}  {len(cols_est)} estaticos")
    print(f"  y      {y.shape}")
    print(f"  primer dia predicho: {fechas[0] + pd.Timedelta(days=1)}  "
          f"ultimo: {fechas[-1] + pd.Timedelta(days=1)}")

    for nombre, arr in [("X_hist", X_hist), ("X_fut", X_fut), ("X_est", X_est)]:
        n_nan = int(np.isnan(arr).sum())
        if n_nan:
            pct = 100 * n_nan / arr.size
            if rellenar_ceros:
                print(f"  {nombre}: {n_nan} NaN ({pct:.2f}%) -> 0 (nulos estructurales)")
                np.nan_to_num(arr, copy=False)
            else:
                print(f"  AVISO {nombre}: {n_nan} NaN ({pct:.2f}%) SIN rellenar -- Keras dara loss=nan")

    meta = {"cols_hist": cols_real, "cols_fut": cols_fut_full, "cols_est": cols_est,
            "ventana_dias": V}
    return X_hist, X_fut, X_est, y, fechas, meta


def dividir_tensores(fechas: np.ndarray):
    """Mascaras booleanas train/val/test sobre el vector de fechas de `construir_tensores`.
    Mismas fronteras que el resto del proyecto, para que los resultados sean comparables."""
    f = pd.to_datetime(fechas)
    fin = pd.Timestamp(MODELO_END) - pd.Timedelta(days=1)
    tr = f <= pd.Timestamp(TRAIN_END)
    va = (f > pd.Timestamp(TRAIN_END)) & (f <= pd.Timestamp(VAL_END))
    te = (f > pd.Timestamp(VAL_END)) & (f <= fin)
    print(f"  train {tr.sum()} dias | validation {va.sum()} | test {te.sum()}")
    return tr, va, te


def dividir_train_val_test_horario(df: pd.DataFrame):
    """Split del dataset horario. Corta por `fecha_pred`, NUNCA por fila: las 24 horas de un dia
    comparten conjunto de informacion y tienen que caer enteras del mismo lado."""
    f = pd.to_datetime(df["fecha_pred"])
    fin_test = pd.Timestamp(MODELO_END) - pd.Timedelta(days=1)
    tr = df[f <= pd.Timestamp(TRAIN_END)]
    va = df[(f > pd.Timestamp(TRAIN_END)) & (f <= pd.Timestamp(VAL_END))]
    te = df[(f > pd.Timestamp(VAL_END)) & (f <= fin_test)]
    return tr, va, te


def dividir_train_val_test(dataset: pd.DataFrame):
    """Split cronologico oficial del equipo. Devuelve (train, validation, test).
    NUNCA aleatorio -- ver docstring del modulo."""
    idx = pd.to_datetime(dataset.index)
    fin_test = pd.Timestamp(MODELO_END) - pd.Timedelta(days=1)
    mask_train = idx <= pd.Timestamp(TRAIN_END)
    mask_val = (idx > pd.Timestamp(TRAIN_END)) & (idx <= pd.Timestamp(VAL_END))
    mask_test = (idx > pd.Timestamp(VAL_END)) & (idx <= fin_test)
    return dataset[mask_train], dataset[mask_val], dataset[mask_test]


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Genera los CSV compartidos del equipo")
    # POR DEFECTO SOLO EL HORARIO (23-ago-2026). Es el formato de trabajo del equipo: conserva
    # la forma horaria de las previsiones -- lo unico que mapea uno a uno contra las 24 salidas
    # -- y se convierte en los tensores del Seq2Seq con un reshape. El diario colapsa cada
    # exogena a mean/min/max, asi que queda como baseline tabular y pasa a ser opcional.
    parser.add_argument("--tensores", action="store_true",
                        help="Guarda tambien los tensores encoder/decoder en .npz "
                             "(X_hist, X_fut, X_est, y, fechas) para el Seq2Seq")
    parser.add_argument("--ventana", type=int, default=7,
                        help="Dias del encoder (por defecto 7 = 168 h, ciclo semanal)")
    parser.add_argument("--variante", default="ambas",
                        choices=VARIANTES + ["ambas"],
                        help="Bloque ntc_*_prev_mw: 'sin_ntc_prev' quita las 6 columnas y "
                             "conserva todos los dias; 'sin_2020' las conserva y descarta los "
                             "dias sin ellas; 'ambas' (defecto) genera las dos matrices para "
                             "comparar")
    parser.add_argument("--sin-excel", action="store_true",
                        help="No genera el .xlsx documentado (solo el CSV)")
    parser.add_argument("--filas-excel", type=int, default=0,
                        help="Filas de la hoja MATRIZ. 0 = TODAS (por defecto). Un numero las "
                             "limita, util para una vista rapida de la estructura")
    parser.add_argument("--con-diario", action="store_true",
                        help="Genera TAMBIEN el dataset diario (1 fila por dia, 24 targets). "
                             "Solo como baseline tabular / compatibilidad con el equipo")
    parser.add_argument("--solo-diario", action="store_true",
                        help="Genera unicamente el diario, sin el horario")
    parser.add_argument("--con-espina", action="store_true",
                        help="Genera tambien espina_horaria (~35 MB, mas lento)")
    parser.add_argument("--con-clima", action="store_true",
                        help="Activa las features de ERA5. OJO: es reanalisis, clima REAL de "
                             "D+1 -- fuga. Usar solo para la ablacion 'prevision perfecta' y "
                             "etiquetar los resultados como tal (queda anotado en el .meta.json)")
    parser.add_argument("--sufijo", default="",
                        help="Etiqueta OPCIONAL para separar lineas de trabajo dentro de una "
                             "misma carpeta. Por defecto vacia: el nombre ya lleva la variante "
                             "(dataset_horario_sin_2020_v01.csv) y la carpeta identifica de "
                             "quien es")
    parser.add_argument("--sobrescribir", action="store_true",
                        help="Reescribe la ultima version en vez de crear una nueva (util al "
                             "iterar sobre un bug, para no dejar 20 ficheros identicos)")
    parser.add_argument("--salida", default=None,
                        help="Carpeta de salida (por defecto, la del script)")
    args = parser.parse_args()

    carpeta = Path(args.salida) if args.salida else Path(__file__).parent
    if args.filas_excel == 0:
        args.filas_excel = None
    comun = dict(sufijo=args.sufijo, sobrescribir=args.sobrescribir)

    hacer_horario = not args.solo_diario
    hacer_diario = args.con_diario or args.solo_diario
    if args.solo_diario and args.con_diario:
        parser.error("--solo-diario ya implica el diario; sobra --con-diario")

    if hacer_horario:
        variantes = VARIANTES if args.variante == "ambas" else [args.variante]
        for v in variantes:
            print(f"\n{'='*78}\nDataset horario -- variante: {v}\n{'='*78}")
            dfh = construir_dataset_horario(incluir_clima=args.con_clima, variante=v)
            trh, vah, teh = dividir_train_val_test_horario(dfh)
            dfh = dfh.copy()
            dfh["split"] = "train"
            dfh.loc[vah.index, "split"] = "validation"
            dfh.loc[teh.index, "split"] = "test"

            # el sufijo lleva la variante: dos matrices que se comparan NO pueden confundirse
            suf = f"{args.sufijo}_{v}" if args.sufijo else v
            ruta_h = _ruta_versionada(carpeta, "dataset_horario", sufijo=suf,
                                      sobrescribir=args.sobrescribir)
            _escribir_con_meta(dfh, ruta_h,
                               f"1 fila por (dia objetivo D+1, hora) -- variante {v}",
                               variante=v, incluir_clima=args.con_clima,
                               filas={"train": len(trh), "validation": len(vah),
                                      "test": len(teh)},
                               dias={"train": int(trh["fecha_pred"].nunique()),
                                     "validation": int(vah["fecha_pred"].nunique()),
                                     "test": int(teh["fecha_pred"].nunique())})
            print(f"  train      {len(trh):>6} filas ({trh['fecha_pred'].nunique()} dias)")
            print(f"  validation {len(vah):>6} filas ({vah['fecha_pred'].nunique()} dias)")
            print(f"  test       {len(teh):>6} filas ({teh['fecha_pred'].nunique()} dias)")
            print(f"  guardado:  {ruta_h.name}")

            if not args.sin_excel:
                ruta_x = ruta_h.with_suffix(".xlsx")
                exportar_excel(dfh, ruta_x, variante=v, max_filas=args.filas_excel)
                print(f"  documentado: {ruta_x.name}")
            REGISTRO_EXCLUSIONES.clear()   # el registro es POR VARIANTE, no acumulado

    if hacer_diario:
        print("\nConstruyendo dataset maestro (diario, para modelado)...")
        dataset = construir_dataset_diario(incluir_clima=args.con_clima)
        train, val, test = dividir_train_val_test(dataset)
        dataset = dataset.copy()
        dataset["split"] = "train"
        dataset.loc[val.index, "split"] = "validation"
        dataset.loc[test.index, "split"] = "test"

        ruta = _ruta_versionada(carpeta, "dataset_diario", **comun)
        _escribir_con_meta(dataset, ruta, "1 fila por dia D, target = 24 horas de D+1",
                           float_format="%.2f", index=True, index_label="fecha",
                           incluir_clima=args.con_clima,
                           dias={"train": len(train), "validation": len(val), "test": len(test)})
        print(f"dataset diario: {dataset.shape[0]} dias x {dataset.shape[1]} columnas")
        print(f"  corte de modelado: ultimo dia predicho = {MODELO_END} "
              f"(datos leidos hasta {DATASET_END})")
        print(f"  train:      {len(train):>5} dias  ({train.index.min()} -> {train.index.max()})")
        print(f"  validation: {len(val):>5} dias  ({val.index.min()} -> {val.index.max()})")
        print(f"  test:       {len(test):>5} dias  ({test.index.min()} -> {test.index.max()})")
        print(f"guardado en: {ruta.name}  (+ {ruta.with_suffix('.meta.json').name})")

    if args.tensores:
        print("\nConstruyendo tensores encoder/decoder...")
        Xh, Xf, Xe, y, fechas, meta = construir_tensores(ventana_dias=args.ventana,
                                                         incluir_clima=args.con_clima)
        tr, va, te = dividir_tensores(fechas)
        ruta_t = _ruta_versionada(carpeta, "tensores", ext=".npz", **comun)
        np.savez_compressed(ruta_t, X_hist=Xh, X_fut=Xf, X_est=Xe, y=y,
                            fechas=fechas.astype("datetime64[D]"),
                            train=tr, val=va, test=te, **{f"cols_{k}": np.array(v, dtype=object)
                                                          for k, v in meta.items()
                                                          if isinstance(v, list)})
        print(f"guardado en: {ruta_t.name}  ({ruta_t.stat().st_size/1e6:.1f} MB)")

    if args.con_espina:
        print("\nConstruyendo espina horaria (para EDA/correlaciones)...")
        espina = construir_espina_horaria()
        ruta_e = _ruta_versionada(carpeta, "espina_horaria", **comun)
        _escribir_con_meta(espina, ruta_e, "union horaria de las tablas nucleo (solo EDA)")
        print(f"espina: {espina.shape[0]} filas x {espina.shape[1]} columnas")
        print(f"rango: {espina['ts'].min()} -> {espina['ts'].max()}")
        print(f"guardado en: {ruta_e.name}")